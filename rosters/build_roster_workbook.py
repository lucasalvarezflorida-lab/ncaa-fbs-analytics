"""Build FBS_Rosters_2026.xlsx: 2025 roster base + 2026 transfer-portal overlay.

Inputs (data/): rosters_2025base.json, portal_2026.json, teams_fbs_2026.json
Output: FBS_Rosters_2026.xlsx — 'All Players' master + one sheet per conference.

Columns: Conference, Team, Jersey #, Name, Position, Class (2025), Height,
Weight, Hometown, 2026 Status.
2026 Status values:
    "Portal -> X"        left this team for X (per CFBD portal feed)
    "Portal -> undecided" entered portal, no destination yet
    "Incoming <- X"      arrived via portal from X (new row, limited fields)
    "2025 Sr - verify"   senior last year; confirm return before quoting
Conference label: CFBD's "American Athletic" is renamed "American" (official
name since July 21, 2025).
"""

import json
import os
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")
COLS = ["Conference", "Team", "Jersey #", "Name", "Position", "Class (2025)",
        "Height", "Weight", "Hometown", "2026 Status"]
WIDTHS = [20, 24, 9, 26, 9, 12, 8, 8, 28, 26]
CLASS_MAP = {1: "Fr", 2: "So", 3: "Jr", 4: "Sr", 5: "5th", 6: "6th"}
RENAME = {"American Athletic": "American"}
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")
OUT_FILL = PatternFill("solid", start_color="FCE4E4")   # outgoing: light red
IN_FILL = PatternFill("solid", start_color="E2EFDA")    # incoming: light green


def norm(s):
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z]", "", s.lower())


def fmt_height(h):
    if not h:
        return ""
    h = int(round(h))
    return f"{h // 12}'{h % 12}\""


def fmt_hometown(p):
    loc = ", ".join(x for x in (p.get("homeCity"), p.get("homeState")) if x)
    c = p.get("homeCountry")
    if c and c not in ("USA", "US", "United States"):
        loc = f"{loc} ({c})" if loc else c
    return loc


def main():
    players = json.load(open(os.path.join(DATA, "rosters_2025base.json")))
    portal = json.load(open(os.path.join(DATA, "portal_2026.json")))
    fbs = {t["school"] for t in json.load(open(os.path.join(DATA, "teams_fbs_2026.json")))}

    for p in players:
        p["conference"] = RENAME.get(p.get("conference"), p.get("conference"))

    by_team_name = {}
    for p in players:
        by_team_name.setdefault((p.get("team"), norm(p.get("firstName", "") + p.get("lastName", ""))), []).append(p)

    out_n = in_n = 0
    incoming_rows = []
    conf_of = {p.get("team"): p.get("conference") for p in players}
    for e in portal:
        nm = norm(e.get("firstName", "") + e.get("lastName", ""))
        dest = e.get("destination")
        # outgoing flag on origin roster
        for p in by_team_name.get((e.get("origin"), nm), []):
            p["status"] = f"Portal -> {dest or 'undecided'}"
            out_n += 1
        # incoming row at destination (FBS only, skip if already on that roster)
        if dest in fbs:
            if by_team_name.get((dest, nm)):
                for p in by_team_name[(dest, nm)]:
                    p["status"] = f"Arrived <- {e.get('origin')}"
            else:
                incoming_rows.append({
                    "team": dest, "conference": conf_of.get(dest),
                    "firstName": e.get("firstName"), "lastName": e.get("lastName"),
                    "position": e.get("position"),
                    "status": f"Incoming <- {e.get('origin')}",
                })
                in_n += 1

    for p in players:
        if not p.get("status") and (p.get("year") or 0) >= 4:
            p["status"] = "2025 Sr - verify"

    players.extend(incoming_rows)
    players.sort(key=lambda p: (p.get("conference") or "~", p.get("team") or "",
                                p.get("lastName") or "", p.get("firstName") or ""))

    def row_for(p):
        return [
            p.get("conference") or "",
            p.get("team") or "",
            p.get("jersey") if p.get("jersey") is not None else "",
            " ".join(x for x in (p.get("firstName"), p.get("lastName")) if x),
            p.get("position") or "",
            CLASS_MAP.get(p.get("year"), p.get("year") or ""),
            fmt_height(p.get("height")),
            p.get("weight") or "",
            fmt_hometown(p),
            p.get("status") or "",
        ]

    def add_sheet(wb, title, plist, skip_conf=False):
        ws = wb.create_sheet(re.sub(r"[\\/?*\[\]:]", "-", title)[:31])
        cols = COLS[1:] if skip_conf else COLS
        widths = WIDTHS[1:] if skip_conf else WIDTHS
        ws.append(cols)
        for c in ws[1]:
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        for p in plist:
            r = row_for(p)
            ws.append(r[1:] if skip_conf else r)
            st = p.get("status") or ""
            fill = OUT_FILL if st.startswith("Portal") else IN_FILL if st.startswith(("Incoming", "Arrived")) else None
            if fill:
                for c in ws[ws.max_row]:
                    c.fill = fill
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = BODY_FONT
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb = Workbook()
    wb.remove(wb.active)

    rm = wb.create_sheet("Read Me")
    rm.column_dimensions["A"].width = 110
    notes = [
        "FBS Rosters - 2026 season prep (built June 2026)",
        "",
        "CONSTRUCTION: 2025 CFBD rosters (CollegeFootballData.com) + 2026 transfer-portal overlay (4,425 entries).",
        "CFBD had no 2026 roster data at build time. Official team athletics sites remain ground truth.",
        "",
        "2026 STATUS COLUMN / ROW COLORS:",
        "  Red rows   'Portal -> X': left this team per CFBD portal feed.",
        "  Green rows 'Incoming <- X' / 'Arrived <- X': portal arrival (limited fields for new rows).",
        "  '2025 Sr - verify': senior last season; confirm return before quoting on air.",
        "",
        "KNOWN GAPS: incoming 2026 freshmen and walk-ons are NOT included (not in CFBD yet).",
        "Class column shows 2025 class standing, not 2026.",
        "Conference labels use 2026 membership; 'American Athletic' renamed 'American' (official since Jul 21, 2025).",
        "",
        "VALIDATION (vs official spring rosters, June 2026):",
        "  Rice: 63/76 official players present (50 from 2025 base, 13 portal incoming); 13 missing = freshmen/walk-ons.",
        "  Memphis: 76/98 present (26 base + 50 portal incoming); 22 missing = freshmen/walk-ons. Heavy churn caught by portal overlay.",
        "  Stale unflagged rows are mostly 2025 seniors (already flagged 'verify').",
        "",
        "FOR ON-AIR USE: for any team featured in an episode, cross-check the official athletics-site roster first.",
        "Refresh: re-run fetch_rosters.py + build_roster_workbook.py once CFBD publishes 2026 rosters.",
    ]
    for i, line in enumerate(notes, 1):
        rm.cell(row=i, column=1, value=line).font = Font(name="Arial", bold=(i == 1))

    add_sheet(wb, "All Players", players)
    confs = sorted({p.get("conference") or "Unknown" for p in players})
    for cf in confs:
        add_sheet(wb, cf, [p for p in players if (p.get("conference") or "Unknown") == cf], skip_conf=True)

    out = os.path.join(HERE, "FBS_Rosters_2026.xlsx")
    wb.save(out)
    print(f"Saved: {len(players)} rows ({out_n} flagged out, {in_n} incoming added), "
          f"{len(confs)} conference sheets: {confs}")


if __name__ == "__main__":
    main()
