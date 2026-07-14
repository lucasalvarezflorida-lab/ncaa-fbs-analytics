import datetime, json, glob, os, pathlib, re, unicodedata
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = str(pathlib.Path(__file__).resolve().parents[1])
D = HERE + "/data"
RAW = D + "/raw_official_2026"
RENAME = {"American Athletic": "American"}
COLS = ["Conference", "Team", "Jersey #", "Name", "Position", "Class", "Height", "Weight", "Hometown", "Notes"]
WIDTHS = [20, 24, 9, 26, 9, 10, 8, 8, 30, 34]
HF = PatternFill("solid", start_color="1F4E78"); HFONT = Font(name="Arial", bold=True, color="FFFFFF")
BFONT = Font(name="Arial"); TFILL = PatternFill("solid", start_color="E2EFDA"); WFILL = PatternFill("solid", start_color="FFF2CC")

def norm(s):
    s = unicodedata.normalize("NFKD", s or "").encode("ascii","ignore").decode()
    s = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b\.?", "", s.lower())
    return re.sub(r"[^a-z]", "", s)

def fmt_h(ft, inch):
    if ft in (None, "", 0): return ""
    return f"{int(ft)}'{int(inch or 0)}\""

def parse_h(s):
    if not s: return ""
    m = re.search(r"(\d)\D+(\d{1,2})", str(s))
    return f"{m.group(1)}'{m.group(2)}\"" if m else str(s)

def to_int(x):
    try: return int(re.sub(r"[^\d]", "", str(x))) if x not in (None, "") else ""
    except Exception: return ""

def normalize(p, src):
    if src == "v2":
        return dict(first=p.get("firstName"), last=p.get("lastName"), jersey=p.get("jerseyNumber"),
            pos=p.get("positionShort"), cls=p.get("academicYearShort") or p.get("academicYearLong"),
            ht=fmt_h(p.get("heightFeet"), p.get("heightInches")), wt=to_int(p.get("weight")), home=p.get("hometown"))
    if src == "rosterxml":
        pi = p.get("playerinfo") or {}
        return dict(first=p.get("firstname"), last=p.get("lastname"), jersey=pi.get("uni"),
            pos=pi.get("pos_short"), cls=pi.get("year"), ht=parse_h(pi.get("height")),
            wt=to_int(pi.get("weight")), home=pi.get("hometown"))
    if src == "wmt":
        return dict(first=p.get("firstName"), last=p.get("lastName"), jersey=p.get("jersey"),
            pos=p.get("pos"), cls=p.get("class"), ht=fmt_h(p.get("hf"), p.get("hi")),
            wt=to_int(p.get("weight")), home=p.get("hometown"))
    return dict(first=p.get("firstName"), last=p.get("lastName"), jersey=p.get("jersey"),
        pos=p.get("pos"), cls=p.get("class"), ht=parse_h(p.get("height")),
        wt=to_int(p.get("weight")), home=p.get("hometown"))

teams = json.load(open(f"{D}/teams_fbs_2026.json"))
conf = {t["school"]: RENAME.get(t["conference"], t["conference"]) for t in teams}
portal = json.load(open(f"{D}/portal_2026.json"))
arrivals = {}
for e in portal:
    if e.get("destination"):
        arrivals[(e["destination"], norm((e.get("firstName") or "")+(e.get("lastName") or "")))] = e.get("origin")

rows, coverage = [], []
for school in sorted(conf):
    fn = f"{RAW}/{school.replace('/','-')}.json"
    d = json.load(open(fn)) if os.path.exists(fn) else None
    if d and d.get("players"):
        src = d.get("source", "v2")
        for p in d["players"]:
            n = normalize(p, src)
            if not (n["first"] or n["last"]): continue
            note = ""
            org = arrivals.get((school, norm((n["first"] or "")+(n["last"] or ""))))
            if org: note = f"Transfer from {org} (2026 portal)"
            cl = str(n["cls"] or "").strip().rstrip(".")
            rows.append({"conference": conf[school], "team": school, "jersey": to_int(n["jersey"]),
                "name": (str(n["first"] or "").strip() + " " + str(n["last"] or "").strip()).strip(),
                "pos": str(n["pos"] or "").strip(), "cls": cl, "ht": n["ht"], "wt": n["wt"],
                "home": str(n["home"] or "").strip(), "note": note,
                "last": str(n["last"] or "").strip(), "first": str(n["first"] or "").strip()})
        coverage.append([school, conf[school], len(d["players"]), d.get("title",""), src, "official site"])
    else:
        # ULM fallback: 2025 base + portal flags
        base = [p for p in json.load(open(f"{D}/rosters_2025base.json")) if p["team"] == school]
        outs = {norm((e.get("firstName") or "")+(e.get("lastName") or "")) for e in portal if e.get("origin") == school}
        kept = 0
        for p in base:
            nm = norm((p.get("firstName") or "")+(p.get("lastName") or ""))
            if nm in outs: continue
            h = p.get("height")
            rows.append({"conference": conf[school], "team": school, "jersey": to_int(p.get("jersey")),
                "name": f"{p.get('firstName','')} {p.get('lastName','')}".strip(),
                "pos": p.get("position") or "", "cls": {1:"Fr",2:"So",3:"Jr",4:"Sr"}.get(p.get("year"), ""),
                "ht": f"{int(h//12)}'{int(h%12)}\"" if h else "", "wt": p.get("weight") or "",
                "home": ", ".join(x for x in (p.get("homeCity"), p.get("homeState")) if x),
                "note": "2025 roster (2026 not yet posted)", "last": p.get("lastName") or "", "first": p.get("firstName") or ""})
            kept += 1
        for (dest, nm), org in arrivals.items():
            if dest != school: continue
            e = next(e for e in portal if e.get("destination") == school and norm((e.get("firstName") or "")+(e.get("lastName") or "")) == nm)
            rows.append({"conference": conf[school], "team": school, "jersey": "", "name": f"{e.get('firstName','')} {e.get('lastName','')}".strip(),
                "pos": e.get("position") or "", "cls": "", "ht": "", "wt": "", "home": "",
                "note": f"Incoming transfer from {org} (2026 portal)", "last": e.get("lastName") or "", "first": e.get("firstName") or ""})
            kept += 1
        coverage.append([school, conf[school], kept, "2025 base + portal (2026 not posted)", "fallback", "CFBD"])

rows.sort(key=lambda r: (r["conference"], r["team"], r["last"], r["first"]))

def add_sheet(wb, title, rs, skip_conf=False):
    ws = wb.create_sheet(re.sub(r"[\\/?*\[\]:]", "-", title)[:31])
    cols = COLS[1:] if skip_conf else COLS
    widths = WIDTHS[1:] if skip_conf else WIDTHS
    ws.append(cols)
    for c in ws[1]:
        c.font = HFONT; c.fill = HF; c.alignment = Alignment(horizontal="center")
    for r in rs:
        v = [r["conference"], r["team"], r["jersey"], r["name"], r["pos"], r["cls"], r["ht"], r["wt"], r["home"], r["note"]]
        ws.append(v[1:] if skip_conf else v)
        if r["note"].startswith(("Transfer", "Incoming")):
            for c in ws[ws.max_row]: c.fill = TFILL
        elif r["note"]:
            for c in ws[ws.max_row]: c.fill = WFILL
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row: c.font = BFONT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

wb = Workbook(); wb.remove(wb.active)
rm = wb.create_sheet("Read Me")
rm.column_dimensions["A"].width = 115
fallback_teams = sorted(c[0] for c in coverage if c[4] == "fallback")
stale_teams = sorted(c[0] for c in coverage if "STALE" in str(c[3] or ""))
fresh_n = len(coverage) - len(fallback_teams) - len(stale_teams)
exceptions = []
if stale_teams:
    exceptions.append(f"STALE (site unavailable on refresh; showing June 10 pull): {', '.join(stale_teams)}.")
if fallback_teams:
    exceptions.append(f"NO 2026 ROSTER POSTED: {', '.join(fallback_teams)} - rows are the 2025 roster minus portal")
    exceptions.append("departures, plus portal arrivals, highlighted yellow/green. Re-pull before taping.")
if not exceptions:
    exceptions.append("None - every team has a live 2026 roster from its official site.")
notes = [
 f"FBS Rosters - 2026 season (pulled from official team athletics sites, refreshed {datetime.date.today():%B %d, %Y})",
 "",
 f"SOURCE: each team's official athletics-site roster (ground truth per project source model) - {fresh_n} of {len(coverage)} teams",
 "have live 2026 rosters pulled fresh this refresh. The bulk pull comes directly from the official sites' own",
 "data endpoints (one request per team).",
 "",
 "EXCEPTIONS: " + exceptions[0],
 *exceptions[1:],
 "",
 "GREEN ROWS: arrived via 2026 transfer portal (cross-referenced vs CFBD portal feed; matching is name-based",
 "and conservative - absence of a note does not mean a player is not a transfer).",
 "",
 "COLUMNS: Class is the class standing shown on the official site (Fr/So/Jr/Sr, R- = redshirt, Gr = grad).",
 "Conference labels = 2026 membership; 'American Athletic' is shown as 'American' (renamed Jul 21, 2025).",
 "See the Coverage sheet for per-team source, roster title, and player count.",
 "",
 "TEAM VIEWER: on each conference tab, pick a team in the yellow dropdown (cell B1) and that team's",
 "roster appears below (requires Excel 365). The All Players tab holds the full filterable data.",
 "",
 "CAVEAT: offseason rosters churn until camp. For any team featured on the show, glance at its official",
 "roster page the day before taping. Scripts in this folder re-pull everything in ~3 minutes.",
]
for i, line in enumerate(notes, 1):
    rm.cell(row=i, column=1, value=line).font = Font(name="Arial", bold=(i == 1))

cv = wb.create_sheet("Coverage")
cv.append(["Team", "Conference", "Players", "Roster Title", "Method", "Source"])
for c in cv[1]:
    c.font = HFONT; c.fill = HF
for row in sorted(coverage):
    cv.append(row)
for i, w in enumerate((24, 20, 9, 34, 11, 14), 1):
    cv.column_dimensions[get_column_letter(i)].width = w
for row in cv.iter_rows(min_row=2):
    for c in row: c.font = BFONT
cv.freeze_panes = "A2"; cv.auto_filter.ref = cv.dimensions


def add_viewer_sheet(wb, conf_name, team_blocks, maxlen):
    # team_blocks: list of (team, start_row_in_AllPlayers, count)
    from openpyxl.worksheet.datavalidation import DataValidation
    ws = wb.create_sheet(re.sub(r"[\\/?*\[\]:]", "-", conf_name)[:31])
    n = len(team_blocks)
    ws["A1"] = "Select team:"; ws["A1"].font = Font(name="Arial", bold=True)
    ws["B1"] = team_blocks[0][0]
    ws["B1"].font = Font(name="Arial", bold=True, color="0000FF")
    ws["B1"].fill = PatternFill("solid", start_color="FFFF00")
    for i, (t, s, c) in enumerate(team_blocks, 1):
        ws.cell(row=i, column=27, value=t)
        ws.cell(row=i, column=28, value=s)
        ws.cell(row=i, column=29, value=c)
    for col in ("AA", "AB", "AC"):
        ws.column_dimensions[col].hidden = True
    dv = DataValidation(type="list", formula1="=$AA$1:$AA$%d" % n, allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv); dv.add(ws["B1"])
    ws["D1"] = '=VLOOKUP($B$1,$AA$1:$AC$%d,3,0)&" players  (%d teams in conference)"' % (n, n)
    ws["D1"].font = Font(name="Arial", italic=True, size=9)
    heads = ["Jersey #", "Name", "Position", "Class", "Height", "Weight", "Hometown", "Notes"]
    for i, h in enumerate(heads, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HFONT; c.fill = HF; c.alignment = Alignment(horizontal="center")
    src_cols = ["C", "D", "E", "F", "G", "H", "I", "J"]
    lookup = "VLOOKUP($B$1,$AA$1:$AC$%d," % n
    for r in range(4, 4 + maxlen):
        for ci, sc in enumerate(src_cols, 1):
            idx = "INDEX('All Players'!$%s:$%s,%s2,0)+ROW()-4)" % (sc, sc, lookup)
            if ci in (1, 6):  # numeric: Jersey, Weight
                inner = 'IF(%s="","",%s)' % (idx, idx)
            else:
                inner = '%s&""' % idx
            f = '=IF(ROW()-3>%s3,0),"",%s)' % (lookup, inner)
            cell = ws.cell(row=r, column=ci, value=f)
            cell.font = BFONT
    for i, w in enumerate([9, 26, 9, 10, 8, 8, 30, 34], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

add_sheet(wb, "All Players", rows)
# compute contiguous team blocks (All Players data starts at row 2, same sort order)
blocks = {}
for i, r in enumerate(rows):
    t = r["team"]
    if t not in blocks: blocks[t] = [i + 2, 0]
    blocks[t][1] += 1
maxlen = max(b[1] for b in blocks.values())
for cf in sorted({r["conference"] for r in rows}):
    tb = [(t, blocks[t][0], blocks[t][1]) for t in sorted({r["team"] for r in rows if r["conference"] == cf})]
    add_viewer_sheet(wb, cf, tb, maxlen)

out_path = HERE + "/FBS_Rosters_2026.xlsx"
try:
    wb.save(out_path)
except PermissionError:
    out_path = HERE + "/FBS_Rosters_2026_refreshed.xlsx"
    wb.save(out_path)
    print("WARNING: FBS_Rosters_2026.xlsx is open/locked - saved as FBS_Rosters_2026_refreshed.xlsx instead")
tn = sum(1 for r in rows if r["note"].startswith(("Transfer", "Incoming")))
print(f"Saved {out_path}: {len(rows)} players / {len(coverage)} teams, {tn} portal arrivals flagged")
print("teams <60 players:", [c[0] for c in coverage if c[2] < 60])
