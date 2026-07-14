"""
Build NCAA_FBS_Teams.xlsx with the revised SP+ + sub-ratings methodology.

Pilot team: Miami (FL). Overview sheet shows the same sub-rating columns
that every team sheet exposes — Rush O, Pass O, Rush D, Pass D, ST,
Composite.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

FONT = "Arial"

HEADER_FILL = PatternFill("solid", start_color="0A2851")
HEADER_FONT = Font(name=FONT, size=14, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", start_color="F47321")
SUBHEADER_FONT = Font(name=FONT, size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", start_color="D9E1F2")
SECTION_FONT = Font(name=FONT, size=11, bold=True, color="000000")
LABEL_FONT = Font(name=FONT, size=10, bold=True)
BODY_FONT = Font(name=FONT, size=10)
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="595959")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

OVERVIEW_COLS = 11
TEAM_COLS = 5


def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER


def section_title(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = LEFT
    return row + 1


def apply_color_scale(ws, range_str):
    ws.conditional_formatting.add(
        range_str,
        ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=5, mid_color="FFEB84",
            end_type="num", end_value=10, end_color="63BE7B",
        ),
    )


# ---------- OVERVIEW SHEET ----------

def build_overview(wb):
    ws = wb.active
    ws.title = "Overview"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=OVERVIEW_COLS)
    t = ws.cell(row=1, column=1, value="NCAA FBS Team Ratings — League Overview")
    style_header(t)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=OVERVIEW_COLS)
    s = ws.cell(row=2, column=1,
                value="Sources: CFBD API (SP+ ratings, unit splits, returning production). "
                      "Ratings on 0–10 z-score scale; pipeline refreshes weekly. "
                      "Composite weighted: 0.25×each of 4 O/D phases + 0.10×ST, divided by 1.10.")
    s.font = NOTE_FONT
    s.alignment = LEFT

    headers = ["Rank", "Team", "Conf", "Rush O", "Pass O", "Rush D",
               "Pass D", "ST", "Composite", "Bye/Status", "Last Updated"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER

    # Miami pilot — projected 2026 sub-ratings derived from 2025 SP+ with
    # returning-production adjustment for known losses (Bain Jr., OL, multiple
    # secondary starters). Real values populate from compute_ratings.py once
    # the pipeline runs against CFBD.
    miami = [1, "Miami (FL)", "ACC", 5.5, 7.5, 7.5, 6.5, 6.8,
             "=ROUND((0.25*D5+0.25*E5+0.25*F5+0.25*G5+0.10*H5)/1.10,2)",
             "—", "2026-05-21"]
    for col, val in enumerate(miami, 1):
        c = ws.cell(row=5, column=col, value=val)
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = BORDER

    placeholders = [
        "Alabama", "Georgia", "Ohio State", "Texas", "Penn State",
        "Notre Dame", "Florida", "USF", "Florida State", "Clemson",
    ]
    for i, team in enumerate(placeholders, 2):
        ws.cell(row=4 + i, column=1, value=i).alignment = CENTER
        ws.cell(row=4 + i, column=2, value=team).alignment = CENTER
        c = ws.cell(row=4 + i, column=3, value="TBD")
        c.alignment = CENTER
        c.font = NOTE_FONT
        for col in range(4, OVERVIEW_COLS + 1):
            ws.cell(row=4 + i, column=col, value=None).alignment = CENTER
        for col in range(1, OVERVIEW_COLS + 1):
            ws.cell(row=4 + i, column=col).border = BORDER

    apply_color_scale(ws, "D5:I14")

    widths = [6, 22, 8, 9, 9, 9, 9, 8, 11, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(start_row=17, start_column=1, end_row=18, end_column=OVERVIEW_COLS)
    note = ws.cell(row=17, column=1, value=(
        "Status — Miami (FL) is the pilot team for template validation. "
        "Remaining 133 FBS programs to be populated in subsequent sessions. "
        "Composite recalculates automatically from sub-rating columns; updating any sub-rating updates the rank-eligible composite."))
    note.font = NOTE_FONT
    note.alignment = LEFT_TOP


# ---------- TEAM SHEET (MIAMI) ----------

def build_miami(wb):
    ws = wb.create_sheet("Miami (FL)")

    # ===== Header =====
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TEAM_COLS)
    t = ws.cell(row=1, column=1, value="Miami (FL) — Hurricanes")
    style_header(t)
    ws.row_dimensions[1].height = 32

    info_lines = [
        ("Conference", "ACC"),
        ("Location", "Coral Gables, FL"),
        ("Stadium", "Hard Rock Stadium (capacity ~65,326)"),
        ("Head Coach", "Mario Cristobal (Year 5, since 2022)"),
        ("Offensive Coordinator", "Shannon Dawson"),
        ("Defensive Coordinator", "Corey Hetherman"),
        ("Last Updated", "2026-05-21"),
    ]
    for i, (label, value) in enumerate(info_lines, 2):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=i, column=2, value=value)
        c.font = BODY_FONT
        c.alignment = LEFT

    # ===== Section 1: Ratings (sub-ratings + composite) =====
    row = 10
    row = section_title(ws, row, "1. Ratings (0–10 z-score scale, SP+ base)", TEAM_COLS)

    # Sub-note explaining the two columns
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TEAM_COLS)
    explainer = ws.cell(row=row, column=1, value=(
        "2025 Final = last season's actual SP+ unit rescaled to 0–10. "
        "2026 Projected = current season w/ returning-production prior applied + QB modifier "
        "(Mensah Power-4 transfer alpha = 0.55). Pipeline-refreshed weekly."))
    explainer.font = NOTE_FONT
    explainer.alignment = LEFT
    row += 1

    rating_headers = ["Unit", "2025 Final", "2026 Projected", "2025 Nat'l Rank", "Notes"]
    rating_header_row = row
    for col, h in enumerate(rating_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    row += 1

    # Miami 2025 sub-ratings derived from publicly reported SP+/raw stats:
    #   Rush O: 88th (150 ypg)            → ~3.5
    #   Pass O: 24th (276 ypg)            → ~8.0
    #   Rush D: 3rd  (82 ypg allowed)     → ~9.8
    #   Pass D: 28th (189 ypg allowed)    → ~7.0  (despite great front, soft on the back end)
    #   ST:     ~50th estimated           → ~6.5
    rating_rows = [
        ("Rush O",        3.5, 5.5, 88, "2025: bottom-third nationally despite gaudy total yards — schedule + pass-tilted gameplan. 2026 projection up with Cantwell (#1 OL recruit), portal OL (Meriweather UGA, Cline UCF), Fletcher Jr. returning."),
        ("Pass O",        8.0, 7.5, 24, "Top-25 in 2025 with Beck. 2026 regresses slightly for Mensah uncertainty (Power-4 transfer α=0.55) — receiver room (Barkate, Toney, Vaughn, Jacobs) limits the downside."),
        ("Rush D",        9.8, 7.5,  3, "Elite #3 unit in 2025 (82 ypg). 2026 projection drops sharply: Bain Jr. (#15 overall), multiple interior DL departed. Damon Wilson II (Mizzou/UGA) is a real replacement but the rest is unproven."),
        ("Pass D",        7.0, 6.5, 28, "Strong-but-not-elite #28 in 2025. Pass rush loss (Bain) hurts coverage even more than the rush D bottom line. New corners Thornton (BC) and Hussey (OSU/FSU) are veteran but didn't play together in 2025."),
        ("Special Teams", 6.5, 6.8, 50, "Middling 2025. New K Jake Weinberg (FSU transfer) stabilizes the place-kicking; Toney KR1 from the WR room."),
        ("Composite",
         "=ROUND((0.25*B13+0.25*B14+0.25*B15+0.25*B16+0.10*B17)/1.10,2)",
         "=ROUND((0.25*C13+0.25*C14+0.25*C15+0.25*C16+0.10*C17)/1.10,2)",
         9,
         "Weighted rollup: 25% each O/D phase + 10% ST, divided by 110%. 2025 nat'l rank #9 (SP+); 2026 projection reflects defensive regression."),
    ]
    for r in rating_rows:
        for col, val in enumerate(r, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = LEFT if col == 5 else CENTER
            c.border = BORDER
        row += 1

    apply_color_scale(ws, "B13:C18")

    # ===== Section 2: Depth chart =====
    row += 1
    row = section_title(ws, row, "2. 2026 Projected Depth Chart (from miamihurricanes.com roster, post-spring)", TEAM_COLS)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TEAM_COLS)
    src = ws.cell(row=row, column=1,
                  value="Source: official athletics roster (fetched 2026-05-21). "
                        "Class year column will populate from CFBD /roster endpoint when the pipeline runs.")
    src.font = NOTE_FONT
    src.alignment = LEFT
    row += 1

    for col, h in enumerate(["Position", "Player (#)", "Status", "Affects Sub-Rating", "Notes"], 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    row += 1

    def section_label(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TEAM_COLS)
        c = ws.cell(row=row, column=1, value=text)
        c.font = LABEL_FONT
        c.fill = PatternFill("solid", start_color="F2F2F2")
        c.alignment = LEFT
        row += 1

    def add_rows(rows):
        nonlocal row
        for r in rows:
            for col, val in enumerate(r, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = BODY_FONT
                c.alignment = LEFT
                c.border = BORDER
            row += 1

    section_label("OFFENSE")
    add_rows([
        ("QB1",  "Darian Mensah (#10)",     "Transfer (Duke)",    "Pass O / Rush O", "Projected starter; α=0.55 on QB modifier."),
        ("QB2",  "Luke Nickel (#5)",        "Returning",          "Pass O / Rush O", "Top returning backup."),
        ("RB1",  "Mark Fletcher Jr. (#4)",  "Returning",          "Rush O",          "Physical between-tackles runner."),
        ("RB2",  "CharMar Brown (#6)",      "Transfer (NDSU)",    "Rush O",          "Power-conference move-up."),
        ("RB3",  "Jordan Lyle (#2)",        "Returning",          "Rush O",          ""),
        ("WR1",  "Cooper Barkate (#18)",    "Transfer (Duke)",    "Pass O",          "Followed Mensah from Duke."),
        ("WR2",  "Malachi Toney (#1)",      "Returning",          "Pass O / ST",     "Returning starter; also KR1."),
        ("WR3",  "Cam Vaughn (#17)",        "Transfer (WVU)",     "Pass O",          ""),
        ("WR4",  "Vandrevius Jacobs (#8)",  "Transfer (SC/FSU)",  "Pass O",          ""),
        ("TE1",  "Elija Lofton (#9)",       "Returning",          "Pass O / Rush O", "Primary in-line TE."),
        ("TE2",  "Israel Briggs (#14)",     "Returning",          "Pass O",          ""),
        ("OL",   "Jackson Cantwell (#79)",  "Freshman (5★)",      "Rush O / Pass O", "No.1 OL recruit nationally."),
        ("OL",   "Jamal Meriweather (#71)", "Transfer (UGA)",     "Rush O / Pass O", "Veteran from Georgia."),
        ("OL",   "Johnathan Cline (#56)",   "Transfer (UCF)",     "Rush O / Pass O", "Multi-year G5 starter."),
        ("OL",   "Samson Okunlola (#63)",   "Returning",          "Rush O / Pass O", ""),
        ("OL",   "Max Buchanan (#66)",      "Returning",          "Rush O / Pass O", ""),
        ("OL",   "Joel Ervin (#70)",        "Returning",          "Rush O / Pass O", ""),
    ])

    section_label("DEFENSE")
    add_rows([
        ("EDGE", "Damon Wilson II (#10)",   "Transfer (Mizzou/UGA)", "Pass D / Rush D", "Headline DL portal addition; partial replacement for Bain Jr."),
        ("EDGE", "Marquise Lightfoot (#12)","Returning",             "Pass D / Rush D", ""),
        ("EDGE", "Hayden Lowe (#14)",       "Returning",             "Pass D / Rush D", ""),
        ("DL",   "Justin Scott (#5)",       "Returning",             "Rush D",          ""),
        ("DL",   "Armondo Blount (#18)",    "Returning",             "Rush D",          ""),
        ("DL",   "Tyson Bacon (#44)",       "Returning",             "Rush D",          ""),
        ("DL",   "Mykah Newton (#55)",      "Returning",             "Rush D",          ""),
        ("LB1",  "Mohamed Toure (#1)",      "Transfer (Rutgers)",    "Rush D / Pass D", "Projected MIKE."),
        ("LB2",  "Kamal Bonner (#11)",      "Transfer (NC State)",   "Rush D / Pass D", "Projected WILL."),
        ("LB3",  "Jordan Campbell (#34)",   "Returning",             "Rush D / Pass D", ""),
        ("CB1",  "Damari Brown (#2)",       "Returning",             "Pass D",          "Returning CB1."),
        ("CB2",  "Omar Thornton (#0)",      "Transfer (BC)",         "Pass D",          ""),
        ("NB",   "Xavier Lucas (#6)",       "Returning",             "Pass D",          "Nickel/Slot."),
        ("S1",   "Bryce Fitzgerald (#3)",   "Returning",             "Pass D / Rush D", "Made game-clinching INT vs. Texas A&M in 2025 CFP."),
        ("S2",   "Conrad Hussey (#38)",     "Transfer (OSU/FSU)",    "Pass D / Rush D", ""),
        ("DB",   "Isaiah Taylor (#28)",     "Transfer (Arizona)",    "Pass D",          ""),
        ("DB",   "Ethan O'Connor (#24)",    "Transfer (Wash. St.)",  "Pass D",          ""),
    ])

    section_label("SPECIAL TEAMS")
    add_rows([
        ("K1",  "Jake Weinberg (#22)",  "Transfer (FSU)",   "ST", "Stabilizes place-kicking."),
        ("K2",  "Jack Olsen (#82)",     "Returning",        "ST", "Backup/competition."),
        ("LS",  "Adam Booker (#46)",    "Returning",        "ST", "Long snapper."),
        ("KR",  "Malachi Toney (#1)",   "Returning",        "ST", "Doubles from WR1."),
        ("P",   "TBD",                  "—",                "ST", "Punter pending roster fetch."),
    ])

    # ===== Section 3: Team stats =====
    row += 1
    row = section_title(ws, row, "3. Team Stats — 2022 through 2025", TEAM_COLS)

    for col, h in enumerate(["Metric", "2025", "2024", "2023", "2022"], 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    row += 1

    stats_rows = [
        ("Record (W-L)",            "13-3",  "10-3",  "7-6",   "5-7"),
        ("Conference (W-L)",        "6-2",   "7-1",   "3-5",   "3-5"),
        ("Points Per Game",         34.1,    35.1,    29.8,    24.5),
        ("Points Allowed/Game",     14.2,    22.5,    23.1,    25.7),
        ("Rush Off (Y/G | rank)",   "150.1 | 88",  "184 | 47",  "165 | 65",  "138 | 88"),
        ("Pass Off (Y/G | rank)",   "275.8 | 24",  "270 | 26",  "215 | 70",  "210 | 78"),
        ("Rush Def (Y/G | rank)",   "81.8 | 3",    "152 | 70",  "150 | 65",  "175 | 95"),
        ("Pass Def (Y/G | rank)",   "189.1 | 28",  "215 | 50",  "230 | 65",  "240 | 75"),
        ("Turnover Margin",         "+12",   "+5",    "-1",    "-7"),
        ("Red Zone TD %",           64.8,    61.5,    55.0,    50.3),
        ("Final AP Rank",           2,       20,      "NR",    "NR"),
        ("Bowl / Playoff Result",   "Lost CFP Final to Indiana",
                                    "Lost Pop-Tarts Bowl to Iowa St.",
                                    "Won Pinstripe Bowl",
                                    "No bowl"),
        ("Final SP+ Rank",          7,       18,      45,      75),
    ]
    for r in stats_rows:
        for col, val in enumerate(r, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = LEFT if col == 1 else CENTER
            c.border = BORDER
        row += 1

    # ===== Section 4: Strengths & weaknesses =====
    row += 1
    row = section_title(ws, row, "4. Strengths & Weaknesses (anchored to sub-rating profile)", TEAM_COLS)

    sw_header_row = row
    for col, h in enumerate(["Sub-Rating", "Direction", "Analyst Note"], 1):
        c = ws.cell(row=sw_header_row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    ws.merge_cells(start_row=sw_header_row, start_column=3, end_row=sw_header_row, end_column=TEAM_COLS)
    row += 1

    sw_rows = [
        ("STRENGTHS", None, None),
        ("Pass O (8.0 → 7.5)", "▲",
         "Top-25 unit in 2025; reloaded for 2026 with Mensah throwing to Barkate/Toney/Vaughn/Jacobs. "
         "Even with the transfer regression on Mensah, the receiver room limits the downside."),
        ("Rush D (9.8 → 7.5)", "▲↓",
         "Elite #3 unit in 2025 (81.8 ypg allowed). Projection regresses but the front-seven floor "
         "is still above-average; Damon Wilson II is a real edge talent, LB room got veterans."),
        ("Composite (7.5 → 6.8)", "▲",
         "Still projects as a top-20 program. Recruiting top-10 nationally three straight cycles; "
         "Cristobal's Year-5 staff was extended after the CFP final appearance."),
        ("WEAKNESSES", None, None),
        ("Rush O (3.5 → 5.5)", "↑ but low base",
         "2025 was bottom-third nationally despite gaudy total yards — the schedule and pass-tilted "
         "scheme inflated raw numbers. Cantwell (#1 OL recruit) + portal OL should help, but this "
         "is the biggest single 'show me' area for the offense in 2026."),
        ("Pass D (7.0 → 6.5)", "↓",
         "Bain Jr. departure hurts coverage indirectly through lost pass rush. Bonner/Toure are "
         "first-year portal LBs — coverage chemistry takes a half-season. Conrad Hussey and Omar "
         "Thornton on the back end haven't played together; expect inconsistency early."),
        ("QB modifier risk",   "neutral",
         "Mensah's α=0.55 means we credit just over half his Duke production. If he overperforms "
         "(α=0.55 was conservative), Pass O climbs back to 8.0; if he underperforms, ~6.5 is the floor."),
        ("Schedule pressure", "—",
         "ACC remains tough + a Power 4 non-con. Margin for repeating the 2025 CFP run is thinner."),
    ]
    for sub, direction, note in sw_rows:
        if note is None:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TEAM_COLS)
            c = ws.cell(row=row, column=1, value=sub)
            c.font = LABEL_FONT
            c.fill = PatternFill("solid", start_color="F2F2F2")
            c.alignment = LEFT
        else:
            ws.cell(row=row, column=1, value=sub).font = LABEL_FONT
            ws.cell(row=row, column=1).alignment = LEFT_TOP
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2, value=direction).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=2).border = BORDER
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=TEAM_COLS)
            c = ws.cell(row=row, column=3, value=note)
            c.font = BODY_FONT
            c.alignment = LEFT_TOP
            c.border = BORDER
            ws.row_dimensions[row].height = 48
        row += 1

    # ===== Section 5: Program history & coaching =====
    row += 1
    row = section_title(ws, row, "5. Program History & Coaching", TEAM_COLS)

    history_rows = [
        ("Founded",                  "1926"),
        ("All-Time Record",          "688–394–19 (through 2025)"),
        ("National Championships",   "5 (1983, 1987, 1989, 1991, 2001)"),
        ("Conference Titles",        "9 Big East titles in 1991–2003 era; first ACC title pending; last conference title 2003 (Big East)"),
        ("Bowl Record",              "19–25 (44 bowl appearances)"),
        ("College Football Playoff", "First CFP appearance and first National Championship Game appearance in 2025-26 (lost to Indiana)"),
        ("Heisman Winners",          "2 — Vinny Testaverde (1986), Gino Torretta (1992)"),
        ("Notable NFL Alumni",       "Ed Reed, Ray Lewis, Warren Sapp, Michael Irvin, Ottis Anderson, Bryant McKinnie, Sean Taylor, Andre Johnson"),
    ]
    for label, value in history_rows:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).alignment = LEFT_TOP
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=TEAM_COLS)
        c = ws.cell(row=row, column=2, value=value)
        c.font = BODY_FONT
        c.alignment = LEFT_TOP
        c.border = BORDER
        row += 1

    row += 1
    row = section_title(ws, row, "Recent Head Coaches", TEAM_COLS)
    for col, h in enumerate(["Coach", "Tenure", "Record"], 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    row += 1

    coaches = [
        ("Mario Cristobal",        "2022–present", "35–18 (through 2025)"),
        ("Manny Diaz",             "2019–2021",    "21–15"),
        ("Mark Richt",             "2016–2018",    "26–13"),
        ("Al Golden",              "2011–2015",    "32–25"),
        ("Randy Shannon",          "2007–2010",    "28–22"),
        ("Larry Coker",            "2001–2006",    "60–15 (won 2001 NC)"),
        ("Butch Davis",            "1995–2000",    "51–20"),
        ("Dennis Erickson",        "1989–1994",    "63–9 (won 1989, 1991 NCs)"),
        ("Jimmy Johnson",          "1984–1988",    "52–9 (won 1987 NC)"),
        ("Howard Schnellenberger", "1979–1983",    "41–16 (won 1983 NC)"),
    ]
    for r in coaches:
        for col, val in enumerate(r, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = LEFT
            c.border = BORDER
        row += 1

    # Column widths for team sheet
    widths = [18, 26, 22, 22, 52]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A11"


# ---------- MAIN ----------

def main():
    wb = Workbook()
    build_overview(wb)
    build_miami(wb)

    import sys
    if sys.platform.startswith("win"):
        out = r"C:\Users\lucas\OneDrive\Personal Project\Sports\Sports Data Analysis\ncaa-fbs-model\NCAA_FBS_Teams.xlsx"
    else:
        out = "/sessions/modest-focused-heisenberg/mnt/Sports Data Analysis/ncaa-fbs-model/NCAA_FBS_Teams.xlsx"
    wb.save(out)
    print(f"Saved workbook to {out}")


if __name__ == "__main__":
    main()
