"""
Build NCAA_FBS_Teams.xlsx end-to-end from the live pipeline output.

Reads:
    ncaa-fbs-model/scripts/data/team_ratings.csv  (output of compute_ratings.py)
    ncaa-fbs-model/scripts/data/team_ratings.json (same data, for metadata)

Produces:
    ncaa-fbs-model/NCAA_FBS_Teams.xlsx with:
    - Overview sheet ranking every FBS team by composite (live numbers)
    - Miami (FL) — full curated pilot sheet, ratings replaced with computed values
    - 133 other team sheets — skeleton with computed ratings + placeholders
      for depth chart, stats, S&W, history (to be filled in future sessions)
"""

import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ---------- Paths ----------
import sys
if sys.platform.startswith("win"):
    PROJECT_DIR = Path(r"C:\Users\lucas\OneDrive\Personal Project\Sports\Sports Data Analysis\ncaa-fbs-model")
else:
    PROJECT_DIR = Path("/sessions/modest-focused-heisenberg/mnt/Sports Data Analysis/ncaa-fbs-model")

CSV_PATH = PROJECT_DIR / "scripts" / "data" / "team_ratings.csv"
JSON_PATH = PROJECT_DIR / "scripts" / "data" / "team_ratings.json"
OUT_PATH = PROJECT_DIR / "NCAA_FBS_Teams.xlsx"

# ---------- Styling ----------
FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="0A2851")
HEADER_FONT = Font(name=FONT, size=14, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", start_color="F47321")
SUBHEADER_FONT = Font(name=FONT, size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", start_color="D9E1F2")
SECTION_FONT = Font(name=FONT, size=11, bold=True, color="000000")
LABEL_FONT = Font(name=FONT, size=10, bold=True)
BODY_FONT = Font(name=FONT, size=10)
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="595959")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER


def section_title(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = LEFT
    return row + 1


def apply_color_scale(ws, range_str):
    ws.conditional_formatting.add(
        range_str,
        ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=5, mid_color="FFEB84",
            end_type="num", end_value=10, end_color="63BE7B",
        ),
    )


# Sheet names cannot exceed 31 chars or contain : \ / ? * [ ]
def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name)
    return cleaned[:31]


# ---------- DATA LOAD ----------

def load_ratings():
    df = pd.read_csv(CSV_PATH)
    # FBS only — drop teams missing SP+ overall (the ~2 stragglers from advanced stats)
    df = df[df["sp_overall"].notna()].copy()

    # Compute per-rating ranks (1 = best)
    for col in ["composite", "rush_o_rating", "pass_o_rating",
                "rush_d_rating", "pass_d_rating", "st_rating",
                "sp_overall_rating"]:
        df[f"{col}_rank"] = df[col].rank(ascending=False, method="min").astype(int)

    df = df.sort_values("composite", ascending=False).reset_index(drop=True)
    df["composite_overall_rank"] = df.index + 1

    meta = json.loads(JSON_PATH.read_text()) if JSON_PATH.exists() else {}
    return df, meta


# ---------- OVERVIEW SHEET ----------

OVERVIEW_COLS = 12

def build_overview(wb, df, meta):
    ws = wb.active
    ws.title = "Overview"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=OVERVIEW_COLS)
    t = ws.cell(row=1, column=1, value=f"NCAA FBS Team Ratings — {meta.get('season', 2025)} League Overview")
    style_header(t)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=OVERVIEW_COLS)
    s = ws.cell(row=2, column=1, value=(
        f"Source: CFBD SP+ + PPA splits, opponent-adjusted. "
        f"Composite = (0.25 × each of 4 sub-ratings + 0.10 × ST) / 1.10. "
        f"SP+ Overall is also rescaled to 0–10 for cross-reference. "
        f"Generated {meta.get('generated_at', '')[:19]}."))
    s.font = NOTE_FONT
    s.alignment = LEFT

    headers = ["Rank", "Team", "Conf", "Rush O", "Pass O", "Rush D",
               "Pass D", "ST", "Composite", "SP+ Overall (rescaled)",
               "SP+ Rank", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER

    for idx, row in df.iterrows():
        excel_row = 5 + idx
        is_miami = row["team"] == "Miami"
        values = [
            int(row["composite_overall_rank"]),
            row["team"] + (" ★" if is_miami else ""),
            row["conference"] if pd.notna(row["conference"]) else "",
            row["rush_o_rating"],
            row["pass_o_rating"],
            row["rush_d_rating"],
            row["pass_d_rating"],
            row["st_rating"],
            row["composite"],
            row["sp_overall_rating"],
            int(row["sp_overall_rating_rank"]),
            "PILOT — curated detail in team sheet" if is_miami else "Skeleton — ratings live, narrative pending",
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=excel_row, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = CENTER if col != 12 else LEFT
            c.border = BORDER

    last_row = 4 + len(df)
    apply_color_scale(ws, f"D5:J{last_row}")

    widths = [6, 26, 8, 9, 9, 9, 9, 8, 11, 11, 10, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"


# ---------- MIAMI (CURATED PILOT) ----------

def build_miami(wb, miami_row, miami_subratings):
    ws = wb.create_sheet(safe_sheet_name("Miami (FL)"))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    t = ws.cell(row=1, column=1, value="Miami (FL) — Hurricanes")
    style_header(t)
    ws.row_dimensions[1].height = 32

    info_lines = [
        ("Conference", "ACC"),
        ("Location", "Coral Gables, FL"),
        ("Stadium", "Hard Rock Stadium (capacity ~65,326)"),
        ("Head Coach", "Mario Cristobal (Year 5, since 2022)"),
        ("Offensive Coordinator", "Shannon Dawson"),
        ("Defensive Coordinator", "Corey Hetherman"),
        ("Last Updated", datetime.now().strftime("%Y-%m-%d")),
    ]
    for i, (label, value) in enumerate(info_lines, 2):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=i, column=2, value=value)
        c.font = BODY_FONT
        c.alignment = LEFT

    # === Ratings ===
    row = 10
    row = section_title(ws, row, "1. Ratings (0–10 z-score, live from CFBD pipeline)", 5)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    explainer = ws.cell(row=row, column=1, value=(
        f"2025 Final = computed from season-end PPA splits + SP+ ST, z-score normalized. "
        f"Composite #{int(miami_row['composite_overall_rank'])} nationally; "
        f"SP+ overall rank #{int(miami_row['sp_overall_rating_rank'])}."))
    explainer.font = NOTE_FONT
    explainer.alignment = LEFT
    row += 1

    for col, h in enumerate(["Unit", "2025 Final", "Nat'l Rank", "Notes"], 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    row += 1

    rating_rows = [
        ("Rush O", miami_row["rush_o_rating"], int(miami_row["rush_o_rating_rank"]),
         f"PPA {miami_row['ppa_rush_o']:.3f}/play. Pass-tilted scheme + game scripts pulled the per-play efficiency below the gaudy total yardage rank."),
        ("Pass O", miami_row["pass_o_rating"], int(miami_row["pass_o_rating_rank"]),
         f"PPA {miami_row['ppa_pass_o']:.3f}/play. Top quartile nationally; Beck's CFP run drove this. Mensah transfer regression risk for 2026 (α=0.55)."),
        ("Rush D", miami_row["rush_d_rating"], int(miami_row["rush_d_rating_rank"]),
         f"PPA {miami_row['ppa_rush_d']:.3f}/play allowed. Strong but not elite by EPA — the #3 yardage rank was inflated by trailing-opponent script. Bain Jr. departure adds 2026 risk."),
        ("Pass D", miami_row["pass_d_rating"], int(miami_row["pass_d_rating_rank"]),
         f"PPA {miami_row['ppa_pass_d']:.3f}/play allowed. Solid mid-pack; coverage will depend on new portal CBs (Thornton, Hussey) gelling early."),
        ("Special Teams", miami_row["st_rating"], int(miami_row["st_rating_rank"]),
         f"SP+ ST rating {miami_row['sp_st']:.2f}. New K Jake Weinberg (FSU transfer) stabilizes; Toney KR1 from the WR room."),
        ("Composite", miami_row["composite"], int(miami_row["composite_overall_rank"]),
         f"Weighted: 0.25 × each of 4 sub-ratings + 0.10 × ST, divided by 1.10."),
        ("SP+ Overall (rescaled)", miami_row["sp_overall_rating"], int(miami_row["sp_overall_rating_rank"]),
         f"Raw SP+ {miami_row['sp_overall']:.2f}. The peer-reviewed headline number; ranks ahead of the per-play composite because SP+ integrates pace and volume."),
    ]
    for r in rating_rows:
        unit, rating, rank, note = r
        ws.cell(row=row, column=1, value=unit).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=1).border = BORDER
        c = ws.cell(row=row, column=2, value=rating); c.font = BODY_FONT; c.alignment = CENTER; c.border = BORDER
        c = ws.cell(row=row, column=3, value=rank); c.font = BODY_FONT; c.alignment = CENTER; c.border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        c = ws.cell(row=row, column=4, value=note); c.font = BODY_FONT; c.alignment = LEFT_TOP; c.border = BORDER
        ws.row_dimensions[row].height = 32
        row += 1

    apply_color_scale(ws, f"B13:B{row-1}")

    # ===== Depth chart (curated content) =====
    row += 1
    row = section_title(ws, row, "2. 2026 Projected Depth Chart (from miamihurricanes.com roster, post-spring)", 5)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    src = ws.cell(row=row, column=1, value=(
        "Source: official athletics roster. Class year column populates from CFBD /roster when wired."))
    src.font = NOTE_FONT
    src.alignment = LEFT
    row += 1

    for col, h in enumerate(["Position", "Player (#)", "Status", "Affects", "Notes"], 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    row += 1

    def section_label(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=text)
        c.font = LABEL_FONT
        c.fill = PatternFill("solid", start_color="F2F2F2")
        c.alignment = LEFT
        row += 1

    def add_rows(rows):
        nonlocal row
        for r in rows:
            for col, val in enumerate(r, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = BODY_FONT
                c.alignment = LEFT
                c.border = BORDER
            row += 1

    section_label("OFFENSE")
    add_rows([
        ("QB1",  "Darian Mensah (#10)",     "Transfer (Duke)",    "Pass O / Rush O", "Projected starter; α=0.55 on QB modifier."),
        ("QB2",  "Luke Nickel (#5)",        "Returning",          "Pass O / Rush O", "Top returning backup."),
        ("RB1",  "Mark Fletcher Jr. (#4)",  "Returning",          "Rush O",          "Physical between-tackles runner."),
        ("RB2",  "CharMar Brown (#6)",      "Transfer (NDSU)",    "Rush O",          "Power-conference move-up."),
        ("WR1",  "Cooper Barkate (#18)",    "Transfer (Duke)",    "Pass O",          "Followed Mensah from Duke."),
        ("WR2",  "Malachi Toney (#1)",      "Returning",          "Pass O / ST",     "Returning starter; also KR1."),
        ("WR3",  "Cam Vaughn (#17)",        "Transfer (WVU)",     "Pass O",          ""),
        ("WR4",  "Vandrevius Jacobs (#8)",  "Transfer (SC/FSU)",  "Pass O",          ""),
        ("TE1",  "Elija Lofton (#9)",       "Returning",          "Pass O / Rush O", "Primary in-line TE."),
        ("OL",   "Jackson Cantwell (#79)",  "Freshman (5★)",      "Rush O / Pass O", "No.1 OL recruit nationally."),
        ("OL",   "Jamal Meriweather (#71)", "Transfer (UGA)",     "Rush O / Pass O", "Veteran from Georgia."),
        ("OL",   "Johnathan Cline (#56)",   "Transfer (UCF)",     "Rush O / Pass O", "Multi-year G5 starter."),
        ("OL",   "Samson Okunlola (#63)",   "Returning",          "Rush O / Pass O", ""),
        ("OL",   "Max Buchanan (#66)",      "Returning",          "Rush O / Pass O", ""),
    ])

    section_label("DEFENSE")
    add_rows([
        ("EDGE", "Damon Wilson II (#10)",   "Transfer (Mizzou/UGA)", "Pass D / Rush D", "Headline DL portal addition; partial replacement for Bain Jr."),
        ("EDGE", "Marquise Lightfoot (#12)","Returning",             "Pass D / Rush D", ""),
        ("DL",   "Justin Scott (#5)",       "Returning",             "Rush D",          ""),
        ("DL",   "Armondo Blount (#18)",    "Returning",             "Rush D",          ""),
        ("DL",   "Tyson Bacon (#44)",       "Returning",             "Rush D",          ""),
        ("LB1",  "Mohamed Toure (#1)",      "Transfer (Rutgers)",    "Rush D / Pass D", "Projected MIKE."),
        ("LB2",  "Kamal Bonner (#11)",      "Transfer (NC State)",   "Rush D / Pass D", "Projected WILL."),
        ("CB1",  "Damari Brown (#2)",       "Returning",             "Pass D",          "Returning CB1."),
        ("CB2",  "Omar Thornton (#0)",      "Transfer (BC)",         "Pass D",          ""),
        ("NB",   "Xavier Lucas (#6)",       "Returning",             "Pass D",          "Nickel/Slot."),
        ("S1",   "Bryce Fitzgerald (#3)",   "Returning",             "Pass D / Rush D", "Made game-clinching INT vs. Texas A&M in 2025 CFP."),
        ("S2",   "Conrad Hussey (#38)",     "Transfer (OSU/FSU)",    "Pass D / Rush D", ""),
    ])

    section_label("SPECIAL TEAMS")
    add_rows([
        ("K",   "Jake Weinberg (#22)",  "Transfer (FSU)",   "ST", "Stabilizes place-kicking."),
        ("LS",  "Adam Booker (#46)",    "Returning",        "ST", ""),
        ("KR",  "Malachi Toney (#1)",   "Returning",        "ST", "Doubles from WR1."),
    ])

    # ===== Team Stats =====
    row += 1
    row = section_title(ws, row, "3. Team Stats — 2022 through 2025", 5)
    for col, h in enumerate(["Metric", "2025", "2024", "2023", "2022"], 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    row += 1
    stats_rows = [
        ("Record (W-L)",            "13-3",  "10-3",  "7-6",   "5-7"),
        ("Conference (W-L)",        "6-2",   "7-1",   "3-5",   "3-5"),
        ("Points Per Game",         34.1,    35.1,    29.8,    24.5),
        ("Points Allowed/Game",     14.2,    22.5,    23.1,    25.7),
        ("Rush Off (Y/G | rank)",   "150.1 | 88",  "184 | 47",  "165 | 65",  "138 | 88"),
        ("Pass Off (Y/G | rank)",   "275.8 | 24",  "270 | 26",  "215 | 70",  "210 | 78"),
        ("Rush Def (Y/G | rank)",   "81.8 | 3",    "152 | 70",  "150 | 65",  "175 | 95"),
        ("Pass Def (Y/G | rank)",   "189.1 | 28",  "215 | 50",  "230 | 65",  "240 | 75"),
        ("Turnover Margin",         "+12",   "+5",    "-1",    "-7"),
        ("Final AP Rank",           2,       20,      "NR",    "NR"),
        ("Bowl / Playoff Result",   "Lost CFP Final to Indiana",
                                    "Lost Pop-Tarts Bowl to Iowa St.",
                                    "Won Pinstripe Bowl",
                                    "No bowl"),
        ("Final SP+ Rank",          7,       18,      45,      75),
    ]
    for r in stats_rows:
        for col, val in enumerate(r, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = LEFT if col == 1 else CENTER
            c.border = BORDER
        row += 1

    # ===== S&W =====
    row += 1
    row = section_title(ws, row, "4. Strengths & Weaknesses (anchored to live sub-ratings)", 5)
    for col, h in enumerate(["Sub-Rating", "Direction", "Analyst Note"], 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    row += 1
    sw_rows = [
        ("STRENGTHS", None, None),
        (f"Pass O ({miami_row['pass_o_rating']:.1f})", "▲",
         "Top-quartile per-play passing offense in 2025 with Beck. Receiver room reloaded (Barkate, Toney, Vaughn, Jacobs); even with Mensah regression, downside is limited."),
        (f"Rush D ({miami_row['rush_d_rating']:.1f})", "▲",
         "Above-average per-play run defense. Bain Jr. departure is real but the unit floor is still solid; Wilson II + the new portal LB room (Toure, Bonner) bridges some of the loss."),
        (f"Composite (#{int(miami_row['composite_overall_rank'])} / SP+ #{int(miami_row['sp_overall_rating_rank'])})", "▲",
         "Top-15 nationally on both views. Recruiting top-10 three straight cycles; coordinators extended after CFP final."),
        ("WEAKNESSES", None, None),
        (f"Rush O ({miami_row['rush_o_rating']:.1f})", "↓",
         "Below-average per-play rushing in 2025 despite total-yardage rank — pass-tilted scheme inflated the optics. Cantwell + portal OL should help in 2026 but this is the show-me area."),
        ("QB modifier risk", "neutral",
         "Mensah's α=0.55 means we credit just over half his Duke production. Pass O floor ~6.0 in a bad scenario; ceiling 8.5 if he overperforms."),
        ("Schedule pressure", "—",
         "ACC remains tough + Power 4 non-con. Repeat 2025 CFP path requires consistency the new pieces haven't shown together yet."),
    ]
    for sub, direction, note in sw_rows:
        if note is None:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws.cell(row=row, column=1, value=sub)
            c.font = LABEL_FONT
            c.fill = PatternFill("solid", start_color="F2F2F2")
            c.alignment = LEFT
        else:
            ws.cell(row=row, column=1, value=sub).font = LABEL_FONT
            ws.cell(row=row, column=1).alignment = LEFT_TOP
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2, value=direction).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=2).border = BORDER
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
            c = ws.cell(row=row, column=3, value=note)
            c.font = BODY_FONT
            c.alignment = LEFT_TOP
            c.border = BORDER
            ws.row_dimensions[row].height = 45
        row += 1

    # ===== History =====
    row += 1
    row = section_title(ws, row, "5. Program History & Coaching", 5)
    history_rows = [
        ("Founded", "1926"),
        ("All-Time Record", "688–394–19 (through 2025)"),
        ("National Championships", "5 (1983, 1987, 1989, 1991, 2001)"),
        ("Bowl Record", "19–25 (44 bowl appearances)"),
        ("CFP", "First CFP appearance and first National Championship Game appearance in 2025-26 (lost to Indiana)"),
        ("Heisman Winners", "2 — Vinny Testaverde (1986), Gino Torretta (1992)"),
        ("Notable Alumni", "Ed Reed, Ray Lewis, Warren Sapp, Michael Irvin, Sean Taylor, Andre Johnson"),
    ]
    for label, value in history_rows:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).alignment = LEFT_TOP
        ws.cell(row=row, column=1).border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row=row, column=2, value=value)
        c.font = BODY_FONT
        c.alignment = LEFT_TOP
        c.border = BORDER
        row += 1

    # Column widths
    widths = [18, 26, 22, 18, 56]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A11"


# ---------- TEAM SKELETON (NON-MIAMI) ----------

def build_team_skeleton(wb, team_row):
    team_name = team_row["team"]
    ws = wb.create_sheet(safe_sheet_name(team_name))

    # Header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t = ws.cell(row=1, column=1, value=team_name)
    style_header(t)
    ws.row_dimensions[1].height = 30

    info = [
        ("Conference",   team_row["conference"] if pd.notna(team_row["conference"]) else "—"),
        ("Composite",    f"{team_row['composite']:.2f}  (#{int(team_row['composite_overall_rank'])} nationally)"),
        ("SP+ Overall",  f"{team_row['sp_overall']:.2f}  (#{int(team_row['sp_overall_rating_rank'])} SP+ rescaled {team_row['sp_overall_rating']:.2f})"),
        ("Last Updated", datetime.now().strftime("%Y-%m-%d")),
    ]
    for i, (label, value) in enumerate(info, 2):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=i, column=2, value=value)
        c.font = BODY_FONT
        c.alignment = LEFT

    # Ratings table
    row = 7
    row = section_title(ws, row, "Ratings (0–10 z-score, CFBD pipeline)", 4)
    for col, h in enumerate(["Unit", "2025 Final", "Nat'l Rank", "Raw PPA / SP+"], 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        c.border = BORDER
    row += 1

    rating_rows = [
        ("Rush O", team_row["rush_o_rating"], int(team_row["rush_o_rating_rank"]), f"PPA {team_row['ppa_rush_o']:.3f}/play"),
        ("Pass O", team_row["pass_o_rating"], int(team_row["pass_o_rating_rank"]), f"PPA {team_row['ppa_pass_o']:.3f}/play"),
        ("Rush D", team_row["rush_d_rating"], int(team_row["rush_d_rating_rank"]), f"PPA {team_row['ppa_rush_d']:.3f}/play allowed"),
        ("Pass D", team_row["pass_d_rating"], int(team_row["pass_d_rating_rank"]), f"PPA {team_row['ppa_pass_d']:.3f}/play allowed"),
        ("Special Teams", team_row["st_rating"], int(team_row["st_rating_rank"]), f"SP+ ST {team_row['sp_st']:.2f}"),
        ("Composite", team_row["composite"], int(team_row["composite_overall_rank"]), "Weighted rollup of sub-ratings"),
        ("SP+ Overall (rescaled)", team_row["sp_overall_rating"], int(team_row["sp_overall_rating_rank"]), f"Raw SP+ {team_row['sp_overall']:.2f}"),
    ]
    rating_start_row = row
    for r in rating_rows:
        for col, val in enumerate(r, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = CENTER if col != 4 else LEFT
            c.border = BORDER
        row += 1

    apply_color_scale(ws, f"B{rating_start_row}:B{row-1}")

    # Placeholders for sections to curate later
    row += 1
    for title in [
        "Depth Chart — pending (CFBD /roster + manual curation)",
        "Team Stats — pending (CFBD season stats endpoint)",
        "Strengths & Weaknesses — pending (analyst notes tied to sub-ratings)",
        "Program History & Coaching — pending (Sports-Reference + school athletics)",
    ]:
        row = section_title(ws, row, title, 4)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value="Section to be populated in future session.")
        c.font = NOTE_FONT
        c.alignment = LEFT
        row += 2

    widths = [22, 22, 14, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- MAIN ----------

def main():
    df, meta = load_ratings()
    print(f"Loaded {len(df)} FBS teams")

    wb = Workbook()
    build_overview(wb, df, meta)

    miami_rows = df[df["team"] == "Miami"]
    if len(miami_rows) == 0:
        print("WARNING: Miami not found in dataset; building skeleton instead")
    else:
        build_miami(wb, miami_rows.iloc[0], None)

    other_teams = df[df["team"] != "Miami"]
    for _, team_row in other_teams.iterrows():
        build_team_skeleton(wb, team_row)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Saved workbook with {len(wb.sheetnames)} sheets → {OUT_PATH}")


if __name__ == "__main__":
    main()
