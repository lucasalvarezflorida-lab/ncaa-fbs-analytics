"""Restructure NCAA_FBS_Teams into conference tabs with team dropdowns.

Replaces the 138 per-team tabs with one tab per conference. Each conference
tab has a team dropdown (cell B1); INDEX/MATCH formulas against three hidden
data sheets display the selected team's info, 2026 schedule, and roster:

  _Teams   - one row per team: conference, FPI decomposition numbers,
             2025 unit ratings/ranks/notes, roster + schedule block pointers
  _Rosters - all players, contiguous per-team blocks (sorted)
  _Sched   - all 2026 games, one row per game per team, contiguous blocks

Data sheets are values-only and rebuilt on every refresh; conference tabs
are formulas and rebuilt too (keeping the current dropdown selection).
"""

from __future__ import annotations

import datetime
import json
import sys
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE / "fpi-decomposition"))

NOTES_JSON = HERE / "team_unit_notes_2025.json"
ROSTER_BOOK = HERE / "rosters" / "FBS_Rosters_2026.xlsx"

ARIAL = Font(name="Arial")
ARIAL_B = Font(name="Arial", bold=True)
WHITE_B = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", start_color="0A2851")
HEAD_FILL = PatternFill("solid", start_color="F47321")
PICK_FILL = PatternFill("solid", start_color="FFFF00")
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="FFFFFF")

CONF_ORDER = ["SEC", "Big Ten", "Big 12", "ACC", "Pac-12", "American",
              "Mountain West", "Sun Belt", "Conference USA", "Mid-American",
              "FBS Independents"]
CONF_ALIAS = {"American Athletic": "American"}

UNITS = ["Rush O", "Pass O", "Rush D", "Pass D", "Special Teams"]

ET = ZoneInfo("America/New_York")


def norm(team: str) -> str:
    from name_mapping import normalize_name
    return normalize_name(team)


# ---------------- data collection ----------------

def read_rosters() -> dict[str, dict]:
    """{team: {conference, players:[...]}} from the roster workbook."""
    wb = openpyxl.load_workbook(ROSTER_BOOK, read_only=True, data_only=True)
    ws = wb["All Players"]
    teams: dict[str, dict] = {}
    for conf, team, jersey, name, pos, cls, ht, wt, home, note in ws.iter_rows(
            min_row=2, values_only=True):
        if not team:
            continue
        t = teams.setdefault(team, {"conference": CONF_ALIAS.get(conf, conf),
                                    "players": []})
        t["players"].append([jersey, name, pos, cls, ht, wt, home, note or ""])
    wb.close()
    return teams


def read_overview_ratings(wb) -> dict[str, dict]:
    """Ratings values per team from the Overview sheet, keyed by norm name."""
    ws = wb["Overview"]
    out = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[1]:
            continue
        rank, team, conf, ro, po, rd, pd_, st, comp, sp_resc, sp_rank = row[:11]
        out[norm(team)] = dict(comp=comp, comp_rank=rank, sp_resc=sp_resc,
                               sp_rank=sp_rank, units=[ro, po, rd, pd_, st])
    return out


def read_unit_notes() -> dict[str, dict]:
    if NOTES_JSON.exists():
        return json.load(open(NOTES_JSON, encoding="utf-8"))
    return {}


LINE_PROVIDERS = ["DraftKings", "Bovada", "ESPN Bet"]  # preference order, per field
HFA = 2.5            # home-field advantage in points for the FPI-implied margin
EDGE_YEL = 6.0       # same-side disagreement threshold (points)
RED_MIN_SPREAD = 3.0 # only call it an upset if a real favorite exists
UNRATED_MARGIN = 24.0  # assumed FPI margin vs FCS / unrated opponents (sim only)


def load_fpi_2026() -> dict:
    """ESPN 2026 preseason FPI from the captured snapshot (norm name -> fpi)."""
    snaps = sorted((HERE / "fpi-decomposition" / "data").glob(
        "fpi_2026_preseason_snapshot_*.json"))
    if not snaps:
        return {}
    out = {}
    for r in json.load(open(snaps[-1], encoding="utf-8")):
        team, v = r.get("team") or r.get("school"), r.get("fpi", r.get("rating"))
        if team and v is not None:
            out[norm(team)] = float(v)
    return out


def fetch_lines(refresh: bool) -> dict:
    """{game_id: {spread, spread_open, spread_text, ou, home_ml, away_ml}}."""
    import cfbd_client as cfbd
    records = cfbd.get("/lines", {"year": 2026, "seasonType": "regular"}, refresh)
    out = {}
    for g in records:
        lines = g.get("lines") or []
        if not lines:
            continue
        ranked = sorted(lines, key=lambda ln: (
            LINE_PROVIDERS.index(ln.get("provider"))
            if ln.get("provider") in LINE_PROVIDERS else 99))

        def first(field):
            for ln in ranked:
                if ln.get(field) is not None:
                    return ln[field]
            return None

        out[g.get("id")] = dict(spread=first("spread"),
                                spread_open=first("spreadOpen"),
                                spread_text=first("formattedSpread"),
                                ou=first("overUnder"),
                                home_ml=first("homeMoneyline"),
                                away_ml=first("awayMoneyline"))
    return out


def _ml_text(v) -> str:
    if v is None:
        return ""
    v = int(v)
    return f"+{v}" if v > 0 else str(v)


def _spread_text(home: str, away: str, spread) -> str:
    if spread is None:
        return ""
    s = float(spread)
    if s == 0:
        return "Pick"
    fav, pts = (home, s) if s < 0 else (away, -s)
    return f"{fav} {pts:g}"


def fetch_games(refresh: bool, fpi: dict[str, dict]) -> list[dict]:
    """All 2026 regular-season games, enriched with lines + FPI edge analytics."""
    import cfbd_client as cfbd
    raw_games = cfbd.get("/games", {"year": 2026, "seasonType": "regular"}, refresh)
    lines = fetch_lines(refresh)

    def pick(g, *names):
        for n in names:
            if g.get(n) is not None:
                return g[n]
        return None

    games = []
    for g in raw_games:
        home = pick(g, "homeTeam", "home_team")
        away = pick(g, "awayTeam", "away_team")
        if not home or not away:
            continue
        raw = pick(g, "startDate", "start_date")
        tbd = pick(g, "startTimeTBD", "start_time_tbd")
        date_s = ""
        if raw:
            dt = datetime.datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
            local = dt.astimezone(ET)
            date_s = f"{local:%a %b %d}"
            if not tbd:
                date_s += f", {local:%I:%M %p ET}".replace(" 0", " ")
        ln = lines.get(pick(g, "id"), {})
        neutral = bool(pick(g, "neutralSite", "neutral_site"))

        hf, af = fpi.get(norm(home)), fpi.get(norm(away))
        model_margin = None  # home perspective, points
        if hf and af:
            model_margin = hf["fpi"] - af["fpi"] + (0 if neutral else HFA)

        spread = ln.get("spread")
        edge = tier = model_side = None
        if model_margin is not None and spread is not None:
            market_margin = -float(spread)  # home perspective
            edge = round(model_margin - market_margin, 1)
            model_side = home if edge > 0 else away
            model_fav = home if model_margin > 0 else away
            market_fav = home if spread < 0 else away
            if model_fav != market_fav and abs(spread) >= RED_MIN_SPREAD:
                tier = "RED"
            elif abs(edge) >= EDGE_YEL:
                tier = "YEL"

        games.append(dict(
            id=pick(g, "id"), wk=pick(g, "week"), date=date_s,
            home=home, away=away, neutral=neutral,
            home_conf=pick(g, "homeConference", "home_conference"),
            away_conf=pick(g, "awayConference", "away_conference"),
            home_class=pick(g, "homeClassification", "home_classification"),
            away_class=pick(g, "awayClassification", "away_classification"),
            venue=pick(g, "venue") or "",
            spread=spread, spread_text=ln.get("spread_text") or "",
            open_text=_spread_text(home, away, ln.get("spread_open")),
            ou=ln.get("ou"), home_ml=ln.get("home_ml"), away_ml=ln.get("away_ml"),
            home_pts=pick(g, "homePoints", "home_points"),
            away_pts=pick(g, "awayPoints", "away_points"),
            completed=bool(pick(g, "completed")),
            model_margin=model_margin, edge=edge, tier=tier,
            model_side=model_side))

    # Market-postmortem flags (market-postmortem/MARKET_POSTMORTEM.md):
    # U-TAIL = total in the season's top decile (2021-25: unders 55.1% on 60+
    # totals, the only spread/total bias to survive FDR). Percentile, not a
    # fixed 60 — the scoring era moves. ML guardrail = RED-alert dogs at +401
    # or longer on the road/neutral render ATS-only (longshot MLs bled -22.9%
    # ROI 2021-25; home dogs, at -1.2%, are exempt).
    posted = sorted(float(g["ou"]) for g in games if g["ou"] is not None)
    tail_thresh = posted[int(0.9 * (len(posted) - 1))] if len(posted) >= 50 else None
    for g in games:
        g["ou_tail"] = (tail_thresh is not None and g["ou"] is not None
                        and float(g["ou"]) >= tail_thresh)
        g["ou_tail_thresh"] = tail_thresh
        dog = dog_ml = None
        ml_guard = False
        if g["tier"] == "RED":
            dog = g["away"] if float(g["spread"]) < 0 else g["home"]
            dog_ml = g["away_ml"] if dog == g["away"] else g["home_ml"]
            ml_guard = (dog_ml is not None and float(dog_ml) >= 401
                        and (g["neutral"] or dog == g["away"]))
        g["dog"], g["dog_ml"], g["ml_guard"] = dog, dog_ml, ml_guard
    if tail_thresh is not None:
        print(f"U-TAIL: totals >= {tail_thresh:g} (top decile of {len(posted)} "
              f"posted; {sum(1 for g in games if g['ou_tail'])} games flagged)")
    return games


def _result_text(g: dict, team_is_home: bool) -> str:
    if not g["completed"] or g["home_pts"] is None:
        return ""
    us = g["home_pts"] if team_is_home else g["away_pts"]
    them = g["away_pts"] if team_is_home else g["home_pts"]
    return f"{'W' if us > them else ('L' if us < them else 'T')} {us}-{them}"


def derive_team_sched(games: list[dict], fpi_rank: dict[str, int]) -> dict[str, list]:
    """Per-team schedule rows for _Sched:
    [wk, date, opp, opp_fpi, ha, spread, open, ou, ml, edge_text, result, venue]"""
    sched: dict[str, list] = {}
    for g in games:
        for team, opp, is_home in ((g["home"], g["away"], True),
                                   (g["away"], g["home"], False)):
            ha = "N" if g["neutral"] else ("vs" if is_home else "at")
            rk = fpi_rank.get(norm(opp))
            edge_text = ""
            if g["edge"] is not None:
                e = g["edge"] if is_home else -g["edge"]
                mark = {"RED": "\U0001F534 ", "YEL": "\U0001F7E1 "}.get(g["tier"], "")
                edge_text = f"{mark}{e:+.1f}"
            ml = g["home_ml"] if is_home else g["away_ml"]
            ou_disp = ""
            if g["ou"] is not None:
                ou_disp = (f"{float(g['ou']):g} ⚑U-TAIL"
                           if g.get("ou_tail") else g["ou"])
            sched.setdefault(team, []).append(
                [g["wk"], g["date"], opp, f"#{rk}" if rk else "n/a", ha,
                 g["spread_text"], g["open_text"], ou_disp, _ml_text(ml),
                 edge_text, _result_text(g, is_home), g["venue"]])
    for team in sched:
        sched[team].sort(key=lambda r: (r[0] or 0))
    return sched


# ---------------- hidden data sheets ----------------

TEAM_COLS = ["Team", "Conference", "FPI Rank", "FPI", "Predicted", "Residual",
             "Resid Rank", "Composite", "Comp Rank", "SP+ (0-10)", "SP+ Rank",
             "RushO", "RushO Rk", "RushO Note", "PassO", "PassO Rk", "PassO Note",
             "RushD", "RushD Rk", "RushD Note", "PassD", "PassD Rk", "PassD Note",
             "ST", "ST Rk", "ST Note",
             "RosterStart", "RosterCount", "SchedStart", "SchedCount",
             "ScoutBaseO", "ScoutTendO", "ScoutBaseD", "ScoutTendD",
             "ScoutStrengths", "ScoutWeaknesses", "DepthStart", "SchemeLabel"]

SCOUT_JSON = HERE / "scouting_top25.json"


def load_scouting() -> dict:
    if SCOUT_JSON.exists():
        return json.load(open(SCOUT_JSON, encoding="utf-8"))
    return {"meta": {"order": []}, "teams": {}}


def build_data_sheets(wb, refresh: bool) -> dict[str, list[str]]:
    """(Re)build _Teams, _Rosters, _Sched. Returns {conference: [teams]}."""
    from analysis import FEATURE_COLS, build_dataset, fit_ols

    rosters = read_rosters()
    ratings = read_overview_ratings(wb)
    notes = read_unit_notes()
    scouting = load_scouting()

    merged, _, _ = build_dataset(2025, refresh=refresh, transfer=True)
    _, fitted = fit_ols(merged, FEATURE_COLS)
    fitted = fitted.sort_values("fpi", ascending=False).reset_index(drop=True)
    fpi = {}
    for i, row in fitted.iterrows():
        fpi[norm(row["team"])] = dict(
            rank=i + 1, fpi=round(row["fpi"], 1), pred=round(row["predicted"], 1),
            resid=round(row["residual"], 1),
            resid_rank=int(fitted["residual"].rank(ascending=False)[i]))

    # overlay ESPN 2026 preseason FPI (captured snapshot) as the live prior:
    # alerts, sim, and schedule opponent ranks all use it; the 2025-fit
    # pred/resid stay as the decomposition context
    fpi26 = load_fpi_2026()
    if fpi26:
        ranked26 = sorted(fpi26.items(), key=lambda kv: -kv[1])
        rank26 = {k: i + 1 for i, (k, _) in enumerate(ranked26)}
        for k, v in fpi26.items():
            e = fpi.setdefault(k, {})
            e["fpi"] = round(v, 1)
            e["rank"] = rank26[k]
            e.setdefault("pred", "n/a")
            e.setdefault("resid", "n/a")
            e.setdefault("resid_rank", "n/a")
        print(f"prior: ESPN 2026 preseason FPI ({len(fpi26)} teams, snapshot)")

    games = fetch_games(refresh, fpi)
    sched = derive_team_sched(games, {k: v["rank"] for k, v in fpi.items()})
    sched_by_norm = {norm(t): g for t, g in sched.items()}

    for name in ("_Teams", "_Rosters", "_Sched", "_DepthGrid"):
        if name in wb.sheetnames:
            del wb[name]
    ws_t = wb.create_sheet("_Teams")
    ws_r = wb.create_sheet("_Rosters")
    ws_s = wb.create_sheet("_Sched")
    ws_g = wb.create_sheet("_DepthGrid")

    ws_t.append(TEAM_COLS)
    ws_r.append(["Team", "#", "Name", "Pos", "Class", "Ht", "Wt", "Hometown", "Notes"])
    ws_s.append(["Team", "Wk", "Date", "Opponent", "Opp FPI", "H/A", "Spread",
                 "Open", "O/U", "ML", "Edge", "Result", "Venue"])
    ws_g.append(["Team", "Pos", "Starter", "Second", "Third", "Fourth"])
    depth_overrides = load_depth_overrides()
    g_row = 2

    conf_teams: dict[str, list[str]] = {}
    r_row, s_row = 2, 2
    for team in sorted(rosters):
        info = rosters[team]
        k = norm(team)
        rat = ratings.get(k, {})
        nt = notes.get(k, {})
        fp = fpi.get(k, {})

        players = info["players"]  # All Players is already sorted by last, first
        team_games = sched_by_norm.get(k, [])

        row = [team, info["conference"],
               fp.get("rank", "n/a"), fp.get("fpi", "n/a"), fp.get("pred", "n/a"),
               fp.get("resid", "n/a"), fp.get("resid_rank", "n/a"),
               rat.get("comp", "n/a"), rat.get("comp_rank", "n/a"),
               rat.get("sp_resc", "n/a"), rat.get("sp_rank", "n/a")]
        units = rat.get("units", ["n/a"] * 5)
        unit_notes = nt.get("notes", [""] * 5)
        unit_ranks = nt.get("ranks", ["n/a"] * 5)
        for u in range(5):
            row += [units[u] if units[u] is not None else "n/a",
                    unit_ranks[u], unit_notes[u]]
        sc = scouting["teams"].get(team, {})
        bullets = lambda xs: "\n".join("• " + x for x in xs) if xs else ""
        grid, scheme_label = build_depth_grid(team, info, sc, depth_overrides)
        row += [r_row, len(players), s_row, len(team_games),
                sc.get("ob", ""), sc.get("ot", ""), sc.get("db", ""),
                sc.get("dt", ""), bullets(sc.get("s")), bullets(sc.get("w")),
                g_row, scheme_label]
        ws_t.append(row)
        for pos, depth4 in grid:
            ws_g.append([team, pos] + depth4)
        g_row += len(grid)

        for p in players:
            ws_r.append([team] + [("" if v is None else v) for v in p])
        r_row += len(players)
        for g in team_games:
            ws_s.append([team] + [("" if v is None else v) for v in g])
        s_row += len(team_games)

        conf_teams.setdefault(info["conference"], []).append(team)

    for ws in (ws_t, ws_r, ws_s, ws_g):
        ws.sheet_state = "hidden"
    team_conf = {t: rosters[t]["conference"] for t in rosters}
    return conf_teams, games, fpi, team_conf


# ---------------- conference tabs ----------------

# fixed layout rows
R_PICK = 1
R_TITLE = 3
R_INFO = 5          # 5..9: Conference, FPI, Residual, Composite, SP+
R_UNIT_HEAD = 11
R_UNIT0 = 12        # 12..16 units
R_SCHED_TITLE = 18
R_SCHED_HEAD = 19
R_SCHED0 = 20
N_SCHED = 14
R_ROSTER_TITLE = R_SCHED0 + N_SCHED + 1   # 35
R_ROSTER_HEAD = R_ROSTER_TITLE + 1        # 36
R_ROSTER0 = R_ROSTER_HEAD + 1             # 37

HC = "AG"  # helper column


def _helper_formulas(ws):
    ws[f"{HC}1"] = "=MATCH($B$1,_Teams!$A:$A,0)"
    ws[f"{HC}2"] = "=INDEX(_Teams!$AA:$AA,$AG$1)"  # roster start
    ws[f"{HC}3"] = "=INDEX(_Teams!$AB:$AB,$AG$1)"  # roster count
    ws[f"{HC}4"] = "=INDEX(_Teams!$AC:$AC,$AG$1)"  # sched start
    ws[f"{HC}5"] = "=INDEX(_Teams!$AD:$AD,$AG$1)"  # sched count
    ws[f"{HC}6"] = "=INDEX(_Teams!$AK:$AK,$AG$1)"  # depth grid start


def _txt(col: str) -> str:
    return f'=IF($AG$1=0,"",INDEX(_Teams!${col}:${col},$AG$1)&"")'


def _num_pair(vcol: str, rcol: str) -> str:
    v = f"INDEX(_Teams!${vcol}:${vcol},$AG$1)"
    r = f"INDEX(_Teams!${rcol}:${rcol},$AG$1)"
    return f'=IF(ISNUMBER({v}),ROUND({v},2)&"  (#"&{r}&")","n/a")'


def build_conference_tab(wb, conf: str, teams: list[str], max_roster: int,
                         stamp: str, prior_pick: str | None):
    import re
    title = re.sub(r"[\\/?*\[\]:]", "-", conf)[:31]
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)

    ws["A1"] = "Select team:"
    ws["A1"].font = ARIAL_B
    pick = prior_pick if prior_pick in teams else teams[0]
    ws["B1"] = pick
    ws["B1"].font = Font(name="Arial", bold=True, size=12, color="0000FF")
    ws["B1"].fill = PICK_FILL
    ws["D1"] = f"{len(teams)} teams — {conf}"
    ws["D1"].font = Font(name="Arial", italic=True, size=9)

    # hidden team list + validation
    for i, t in enumerate(teams, 1):
        ws.cell(row=i, column=34, value=t)  # AH
    dv = DataValidation(type="list", formula1=f"=$AH$1:$AH${len(teams)}",
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws["B1"])
    _helper_formulas(ws)
    for col in ("AG", "AH"):
        ws.column_dimensions[col].hidden = True

    # title bar
    ws.cell(row=R_TITLE, column=1, value="=$B$1").font = TITLE_FONT
    for c in range(1, 9):
        ws.cell(row=R_TITLE, column=c).fill = TITLE_FILL

    info = [
        ("Conference", _txt("B")),
        ("ESPN FPI (2026 preseason)", _num_pair("D", "C")),
        ("Residual (2025 decomposition)", _num_pair("F", "G")),
        ("Composite (0-10)", _num_pair("H", "I")),
        ("SP+ (rescaled 0-10)", _num_pair("J", "K")),
    ]
    for i, (label, f) in enumerate(info):
        ws.cell(row=R_INFO + i, column=1, value=label).font = ARIAL_B
        c = ws.cell(row=R_INFO + i, column=2, value=f)
        c.font = ARIAL

    # unit ratings
    for j, h in enumerate(["Unit", "Rating", "Nat'l Rank", "Raw"], 1):
        c = ws.cell(row=R_UNIT_HEAD, column=j, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
    unit_cols = [("L", "M", "N"), ("O", "P", "Q"), ("R", "S", "T"),
                 ("U", "V", "W"), ("X", "Y", "Z")]
    for i, (u, (vc, rc, nc)) in enumerate(zip(UNITS, unit_cols)):
        r = R_UNIT0 + i
        ws.cell(row=r, column=1, value=u).font = ARIAL_B
        ws.cell(row=r, column=2, value=f'=IF(ISNUMBER(INDEX(_Teams!${vc}:${vc},$AG$1)),INDEX(_Teams!${vc}:${vc},$AG$1),"n/a")').font = ARIAL
        ws.cell(row=r, column=3, value=f'=INDEX(_Teams!${rc}:${rc},$AG$1)&""').font = ARIAL
        ws.cell(row=r, column=4, value=f'=INDEX(_Teams!${nc}:${nc},$AG$1)&""').font = ARIAL

    # schedule
    from openpyxl.formatting.rule import FormulaRule

    ws.cell(row=R_SCHED_TITLE, column=1, value="2026 Schedule").font = ARIAL_B
    note = ws.cell(row=R_SCHED_TITLE, column=4,
                   value="Opp FPI = opponent's 2025 final FPI rank. Lines: DraftKings/Bovada via CFBD as of refresh. "
                         "Edge = FPI-implied margin minus market margin for THIS team; "
                         "\U0001F534 model likes the dog outright, \U0001F7E1 6+ pt disagreement. "
                         "⚑U-TAIL = total in the season's top decile (2021-25 post-mortem: unders 55%).")
    note.font = Font(name="Arial", italic=True, size=9)
    for j, h in enumerate(["Wk", "Date", "Opponent", "Opp FPI", "H/A", "Spread",
                           "Open", "O/U", "ML", "Edge", "Result", "Venue"], 1):
        c = ws.cell(row=R_SCHED_HEAD, column=j, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
    for i in range(N_SCHED):
        r = R_SCHED0 + i
        guard = f"ROW()-{R_SCHED0 - 1}>$AG$5"
        ws.cell(row=r, column=1, value=f'=IF({guard},"",INDEX(_Sched!$B:$B,$AG$4+ROW()-{R_SCHED0}))').font = ARIAL
        for j, col in ((2, "C"), (3, "D"), (4, "E"), (5, "F"), (6, "G"),
                       (7, "H"), (8, "I"), (9, "J"), (10, "K"), (11, "L"),
                       (12, "M")):
            ws.cell(row=r, column=j, value=f'=IF({guard},"",INDEX(_Sched!${col}:${col},$AG$4+ROW()-{R_SCHED0})&"")').font = ARIAL

    rng = f"A{R_SCHED0}:L{R_SCHED0 + N_SCHED - 1}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("\U0001F534",$J{R_SCHED0}))'],
        fill=PatternFill("solid", start_color="FFC7CE"), stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("\U0001F7E1",$J{R_SCHED0}))'],
        fill=PatternFill("solid", start_color="FFEB9C"), stopIfTrue=True))

    # roster
    ws.cell(row=R_ROSTER_TITLE, column=1,
            value=f"2026 Roster — official athletics site (refreshed {stamp})").font = ARIAL_B
    for j, h in enumerate(["#", "Name", "Pos", "Class", "Ht", "Wt", "Hometown", "Notes"], 1):
        c = ws.cell(row=R_ROSTER_HEAD, column=j, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center")
    cols = ["B", "C", "D", "E", "F", "G", "H", "I"]  # _Rosters data cols
    for i in range(max_roster):
        r = R_ROSTER0 + i
        guard = f"ROW()-{R_ROSTER0 - 1}>$AG$3"
        for j, col in enumerate(cols, 1):
            idx = f"INDEX(_Rosters!${col}:${col},$AG$2+ROW()-{R_ROSTER0})"
            if j in (1, 6):  # jersey (B), weight (G): numeric, avoid 0-for-blank
                f = f'=IF({guard},"",IF({idx}="","",{idx}))'
            else:
                f = f'=IF({guard},"",{idx}&"")'
            ws.cell(row=r, column=j, value=f).font = ARIAL

    # ----- right-side panel: strengths/weaknesses + scheme depth chart (N..R) -----
    from openpyxl.styles import Alignment as _Al

    def _panel_header(row, txt, color):
        c = ws.cell(row=row, column=14, value=txt)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=color)
        for cc in range(15, 19):
            ws.cell(row=row, column=cc).fill = PatternFill("solid", start_color=color)

    _panel_header(1, "STRENGTHS", "2E6B4F")
    ws.merge_cells("N2:R5")
    sc = ws["N2"]
    sc.value = '=IF($AG$1=0,"",INDEX(_Teams!$AI:$AI,$AG$1)&"")'
    sc.font = ARIAL
    sc.alignment = _Al(wrap_text=True, vertical="top")
    _panel_header(6, "WEAKNESSES", "A3332C")
    ws.merge_cells("N7:R10")
    wc = ws["N7"]
    wc.value = '=IF($AG$1=0,"",INDEX(_Teams!$AJ:$AJ,$AG$1)&"")'
    wc.font = ARIAL
    wc.alignment = _Al(wrap_text=True, vertical="top")

    ws.merge_cells("N12:R12")
    dc = ws["N12"]
    dc.value = "DEPTH CHART"
    dc.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    for cc in range(14, 19):
        ws.cell(row=12, column=cc).fill = TITLE_FILL
    for j, h in enumerate(["Pos", "Starter", "2nd", "3rd", "4th"], 14):
        c = ws.cell(row=13, column=j, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
    R_DC0 = 14
    N_GRID = N_GRID_ROWS  # 3 section bands + 13 off + 12 def + 7 ST
    BAND_AT = {0, 14, 27}
    BAND_F = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    OFF_F = Font(name="Arial", bold=True, size=10, color="2E6B4F")
    DEF_F = Font(name="Arial", bold=True, size=10, color="A3332C")
    ST_F = Font(name="Arial", bold=True, size=10, color="5C6B7E")
    for i in range(N_GRID):
        r = R_DC0 + i
        band = i in BAND_AT
        pf = BAND_F if band else OFF_F if i < 14 else DEF_F if i < 27 else ST_F
        ws.cell(row=r, column=14,
                value=f'=IF($AG$1=0,"",INDEX(_DepthGrid!$B:$B,$AG$6+ROW()-{R_DC0})&"")').font = pf
        for j, col in ((15, "C"), (16, "D"), (17, "E"), (18, "F")):
            c = ws.cell(row=r, column=j,
                        value=f'=IF($AG$1=0,"",INDEX(_DepthGrid!${col}:${col},$AG$6+ROW()-{R_DC0})&"")')
            c.font = BAND_F if band else Font(name="Arial", size=10)
        if band:
            for cc in range(14, 19):
                ws.cell(row=r, column=cc).fill = PatternFill("solid", start_color="35507A")
    note = ws.cell(row=R_DC0 + N_GRID, column=14,
                   value="Source: OurLads.com curated charts where captured (date in section bands; local use only) — "
                         "'(projected)' bands = our roster-seniority fallback with depth_overrides.csv pins. "
                         "• = transfer/portal, ✓ = pinned.")
    note.font = Font(name="Arial", italic=True, size=8.5, color="5C6B7E")
    ws.merge_cells(start_row=R_DC0 + N_GRID, start_column=14,
                   end_row=R_DC0 + N_GRID + 2, end_column=18)
    note.alignment = _Al(wrap_text=True, vertical="top")

    widths = [22, 30, 26, 9, 8, 16, 30, 34, 28, 8, 8, 8, 2, 6, 24, 22, 20, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


# ---------------- depth charts ----------------

DEPTH_OVERRIDES = HERE / "depth_overrides.csv"
DEPTH_GROUPS = ["QB", "RB", "WR", "TE", "OL", "DL/EDGE", "LB", "DB", "Specialists"]
DEPTH_CAPS = {"QB": 3, "RB": 4, "WR": 6, "TE": 3, "OL": 7, "DL/EDGE": 7,
              "LB": 5, "DB": 8, "Specialists": 3}
_POS_MAP = {
    "QB": "QB",
    "RB": "RB", "FB": "RB", "HB": "RB", "TB": "RB", "AB": "RB",
    "WR": "WR", "ATH": "WR", "SLOT": "WR",
    "TE": "TE", "H": "TE",
    "OL": "OL", "OT": "OL", "OG": "OL", "G": "OL", "C": "OL", "T": "OL",
    "IOL": "OL", "OC": "OL",
    "DL": "DL/EDGE", "DT": "DL/EDGE", "DE": "DL/EDGE", "NT": "DL/EDGE",
    "NG": "DL/EDGE", "EDGE": "DL/EDGE", "RUSH": "DL/EDGE", "JACK": "DL/EDGE",
    "LB": "LB", "ILB": "LB", "OLB": "LB", "MLB": "LB", "WLB": "LB",
    "SLB": "LB", "MIKE": "LB", "WILL": "LB", "SAM": "LB",
    "DB": "DB", "CB": "DB", "S": "DB", "FS": "DB", "SS": "DB",
    "NICKEL": "DB", "NB": "DB", "STAR": "DB", "ROVER": "DB",
    "K": "Specialists", "P": "Specialists", "PK": "Specialists",
    "LS": "Specialists", "K/P": "Specialists", "KO": "Specialists",
}
_CLASS_RANK = {"GR": 0, "GRAD": 0, "6TH": 0, "SR": 1, "R-SR": 1, "RS-SR": 1,
               "5TH": 1, "JR": 2, "R-JR": 2, "RS-JR": 2, "SO": 3, "R-SO": 3,
               "RS-SO": 3, "FR": 4, "R-FR": 4, "RS-FR": 4}


_POS_FULL = {
    "QUARTERBACK": "QB", "RUNNING BACK": "RB", "FULLBACK": "RB", "HALFBACK": "RB",
    "WIDE RECEIVER": "WR", "ATHLETE": "WR", "TIGHT END": "TE",
    "OFFENSIVE LINE": "OL", "OFFENSIVE LINEMAN": "OL", "OFFENSIVE TACKLE": "OL",
    "OFFENSIVE GUARD": "OL", "CENTER": "OL",
    "DEFENSIVE LINE": "DL/EDGE", "DEFENSIVE LINEMAN": "DL/EDGE",
    "DEFENSIVE END": "DL/EDGE", "DEFENSIVE TACKLE": "DL/EDGE",
    "EDGE RUSHER": "DL/EDGE", "NOSE TACKLE": "DL/EDGE",
    "LINEBACKER": "LB", "INSIDE LINEBACKER": "LB", "OUTSIDE LINEBACKER": "LB",
    "DEFENSIVE BACK": "DB", "CORNERBACK": "DB", "SAFETY": "DB", "NICKELBACK": "DB",
    "KICKER": "Specialists", "PUNTER": "Specialists", "PLACE KICKER": "Specialists",
    "LONG SNAPPER": "Specialists", "LONGSNAPPER": "Specialists",
    "DEEP SNAPPER": "Specialists", "PUNTER/KICKER": "Specialists",
}


def _group_of(pos) -> str | None:
    p = str(pos or "").strip().upper().replace(".", "")
    return (_POS_MAP.get(p) or _POS_FULL.get(p)
            or _POS_MAP.get(p.split("/")[0]) or _POS_FULL.get(p.split("/")[0]))


def _cls_rank(cls) -> int:
    return _CLASS_RANK.get(str(cls or "").strip().upper(), 3)


def load_depth_overrides() -> dict:
    """{(norm team, group): [norm player names in pinned order]}"""
    import csv
    out: dict = {}
    if not DEPTH_OVERRIDES.exists():
        return out
    with open(DEPTH_OVERRIDES, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            key = (norm(row["team"]), row["group"].strip())
            out.setdefault(key, []).append(
                (int(row["rank"]), norm(row["player"])))
    return {k: [p for _, p in sorted(v)] for k, v in out.items()}


# --- scheme-customized positional depth grid (conference-tab panel) ---

OFF_TEMPLATES = {
    "11": ["QB", "RB", "WR", "WR", "WR", "TE", "LT", "LG", "C", "RG", "RT"],
    "12": ["QB", "RB", "WR", "WR", "TE", "TE", "LT", "LG", "C", "RG", "RT"],
    "21": ["QB", "RB", "FB", "WR", "WR", "TE", "LT", "LG", "C", "RG", "RT"],
    "10": ["QB", "RB", "WR", "WR", "WR", "WR", "LT", "LG", "C", "RG", "RT"],
    "flex": ["QB", "FB", "AB", "AB", "WR", "WR", "LT", "LG", "C", "RG", "RT"],
}
DEF_TEMPLATES = {
    "425": ["DE", "DT", "DT", "DE", "LB", "LB", "CB", "CB", "NB", "S", "S"],
    "43": ["DE", "DT", "DT", "DE", "WLB", "MLB", "SLB", "CB", "CB", "SS", "FS"],
    "34": ["DE", "NT", "DE", "OLB", "ILB", "ILB", "OLB", "CB", "CB", "SS", "FS"],
    "335": ["DE", "NT", "DE", "LB", "LB", "LB", "CB", "CB", "NB", "S", "S"],
}
ST_ROWS = ["K", "P", "LS", "KR", "PR"]
OFF_LABELS = {"11": "11 personnel (3WR-1TE)", "12": "12 personnel (2WR-2TE)",
              "21": "21 personnel (FB)", "10": "10 personnel (4WR)",
              "flex": "flexbone option"}
DEF_LABELS = {"425": "4-2-5", "43": "4-3", "34": "3-4", "335": "3-3-5"}


def _off_scheme(card: dict) -> str:
    t = (card.get("ob", "") + " " + card.get("ot", "")).lower()
    if "flexbone" in t or "triple option" in t:
        return "flex"
    if "21 personnel" in t or "12/21" in t:
        return "21"
    if "run-and-shoot" in t or "four-wide" in t or "10 personnel" in t or "empties the set" in t:
        return "10"
    if ("12 personnel" in t or "12-personnel" in t or "12/13" in t
            or "multi-te" in t or "te-heavy" in t or "te-friendly" in t
            or "double-tight" in t):
        return "12"
    return "11"


def _def_scheme(card: dict) -> str:
    t = (card.get("db", "") + " " + card.get("dt", "")).lower()
    if "3-3-5" in t or "tite" in t or "3-3" in t:
        return "335"
    if "3-4" in t:
        return "34"
    if "4-3" in t:
        return "43"
    return "425"


def _sub(players, keyset):
    """players whose listed position exactly matches the keyset (upper)."""
    return [p for p in players
            if p["pos"].upper().replace(".", "").strip() in keyset]


def _fmt_depth(p, pins):
    if p is None:
        return ""
    pin = "✓ " if norm(p["name"]) in pins else ""
    return f"{pin}{p['name']} ({p['cls']})" + (" •" if p["portal"] else "")


def _fill_rows(labels, pools, pins):
    """rows: label -> ordered candidates. pools: {label: [players]} pre-split;
    labels sharing a pool are filled column-major."""
    from collections import Counter
    label_ct = Counter(labels)
    seen_idx = Counter()
    grid = []
    for lab in labels:
        pool = pools.get(lab, [])
        k = seen_idx[lab]
        n = label_ct[lab]
        row = [pool[d * n + k] if d * n + k < len(pool) else None for d in range(4)]
        seen_idx[lab] += 1
        grid.append((lab, [_fmt_depth(p, pins) for p in row]))
    return grid


OURLADS_JSON = HERE / "ourlads_depth.json"
GRID_OFF, GRID_DEF, GRID_ST = 13, 12, 7  # fixed section sizes (padded/trimmed)
N_GRID_ROWS = 3 + GRID_OFF + GRID_DEF + GRID_ST  # 35 incl. 3 section bands

_OURLADS_CACHE = None


def load_ourlads() -> dict:
    global _OURLADS_CACHE
    if _OURLADS_CACHE is None:
        if OURLADS_JSON.exists():
            _OURLADS_CACHE = json.load(open(OURLADS_JSON, encoding="utf-8"))
        else:
            _OURLADS_CACHE = {"teams": {}, "captured": ""}
    return _OURLADS_CACHE


_OL_ST = {"PT", "PK", "KO", "LS", "H", "PR", "KR", "P", "K"}
_OL_DEF_PREFIX = ("DE", "DT", "NT", "LB", "CB", "S", "FS", "SS", "NB", "LDE",
                  "RDE", "LDT", "RDT", "WLB", "MLB", "SLB", "LCB", "RCB", "ILB",
                  "OLB", "EDGE", "RUSH", "STAR", "ROVER", "BANDIT", "JACK", "DS")
import re as _re
_CLS_RE = _re.compile(r"\s+((?:RS\s+)?(?:GR|SR|JR|SO|FR)(?:/TR)?|GR/TR|TR)$", _re.I)


def _ol_fmt(raw: str) -> str:
    """'Sayin, Julian RS SO' / 'Martin, Justyn RS SR/TR' -> display string."""
    s = raw.strip()
    cls, transfer = "", "/TR" in s.upper()
    m = _CLS_RE.search(s)
    if m:
        cls = m.group(1).upper().replace("/TR", "").strip()
        s = s[: m.start()].strip()
    if "," in s:
        last, first = s.split(",", 1)
        s = f"{first.strip()} {last.strip()}"
    out = s + (f" ({cls})" if cls else "")
    return out + (" •" if transfer else "")


def _ourlads_sections(entry: dict):
    off, deff, st = [], [], []
    for r in entry.get("rows", []):
        pos = r["pos"].strip().upper()
        players = [_ol_fmt(p) for p in r.get("players", [])[:4]]
        players += [""] * (4 - len(players))
        row = (r["pos"].strip(), players)
        if pos in _OL_ST:
            st.append(row)
        elif pos.startswith(_OL_DEF_PREFIX):
            deff.append(row)
        else:
            off.append(row)
    return off, deff, st


def _pad(section, n):
    section = section[:n]
    return section + [("", ["", "", "", ""])] * (n - len(section))


def build_depth_grid(team: str, info: dict, card: dict, overrides: dict):
    """Fixed 35 rows (3 bands + 13 off + 12 def + 7 ST): (pos, [p1..p4]).
    OurLads curated chart wins when captured; else roster heuristic + pins."""
    ol = load_ourlads()
    entry = ol.get("teams", {}).get(team)
    if entry and entry.get("rows"):
        off, deff, st = _ourlads_sections(entry)
        upd = entry.get("updated", "").split(" ")[0]
        src = f" (OurLads {upd})" if upd else " (OurLads)"
        off_lab = (entry.get("off_scheme") or "base").strip()
        def_lab = (entry.get("def_scheme") or "base").strip()
        rows = ([("", [f"OFFENSE — {off_lab}{src}", "", "", ""])] + _pad(off, GRID_OFF)
                + [("", [f"DEFENSE — {def_lab}{src}", "", "", ""])] + _pad(deff, GRID_DEF)
                + [("", [f"SPECIAL TEAMS{src}", "", "", ""])] + _pad(st, GRID_ST))
        label = f"{off_lab} · {def_lab} (OurLads)"
        return rows, label
    return _heuristic_depth_grid(team, info, card, overrides)


def _heuristic_depth_grid(team: str, info: dict, card: dict, overrides: dict):
    """Roster-derived fallback: scheme template + seniority + pins."""
    groups: dict[str, list] = {g: [] for g in DEPTH_GROUPS}
    for jersey, name, pos, cls, ht, wt, home, note in info["players"]:
        g = _group_of(pos)
        if not g or not name:
            continue
        groups[g].append(dict(name=name, pos=str(pos or "").strip(),
                              cls=str(cls or "").strip(),
                              jersey=jersey if isinstance(jersey, int) else 999,
                              portal=str(note or "").startswith(("Transfer", "Incoming"))))
    all_pins = set()
    for g in DEPTH_GROUPS:
        pins = overrides.get((norm(team), g), [])
        all_pins.update(pins)

        def key(p, pins=pins):
            pn = norm(p["name"])
            return (pins.index(pn) if pn in pins else 99,
                    _cls_rank(p["cls"]), p["jersey"], p["name"])
        groups[g].sort(key=key)

    off_k, def_k = _off_scheme(card), _def_scheme(card)
    off, deff = OFF_TEMPLATES[off_k], DEF_TEMPLATES[def_k]

    # --- offensive line: use OT/G/C detail when the roster has it, else stride ---
    ol = groups["OL"]
    ots = _sub(ol, {"OT", "T", "OFFENSIVE TACKLE", "TACKLE"})
    gs = _sub(ol, {"OG", "G", "OFFENSIVE GUARD", "GUARD"})
    cs = _sub(ol, {"OC", "C", "CENTER"})
    if len(ots) >= 2 and len(gs) >= 2:
        p_lt, p_rt = ots[0::2], ots[1::2]
        p_lg, p_rg = gs[0::2], gs[1::2]
        p_c = cs or gs
    else:
        p_lt, p_lg, p_c, p_rg, p_rt = (ol[0::5], ol[1::5], ol[2::5],
                                       ol[3::5], ol[4::5])
    # --- defensive line ---
    dl = groups["DL/EDGE"]
    edges = _sub(dl, {"DE", "EDGE", "RUSH", "JACK", "DEFENSIVE END", "END", "OLB/DE"})
    inter = _sub(dl, {"DT", "NT", "NG", "DEFENSIVE TACKLE", "NOSE TACKLE",
                      "NOSE GUARD", "DEFENSIVE LINE", "DEFENSIVE LINEMAN", "DL"})
    inter = [p for p in inter if p not in edges]
    if len(edges) < 2 or len(inter) < 2:
        edges, inter = dl[0::2], dl[1::2]
    # --- secondary ---
    dbs = groups["DB"]
    cbs = _sub(dbs, {"CB", "CORNERBACK", "CORNER"})
    safs = _sub(dbs, {"S", "FS", "SS", "SAF", "SAFETY", "FREE SAFETY", "STRONG SAFETY"})
    nbs = _sub(dbs, {"NICKEL", "NB", "STAR", "ROVER", "NICKELBACK"})
    if len(cbs) < 2 or len(safs) < 2:
        cbs, safs = dbs[0::2], dbs[1::2]
    if not nbs:
        nbs = cbs[2:] + safs[2:]
    # --- LB splits by template ---
    lbs = groups["LB"]
    # --- specialists ---
    st = groups["Specialists"]
    ks = _sub(st, {"K", "PK", "KICKER", "PLACE KICKER", "K/P"}) or st
    ps = _sub(st, {"P", "PUNTER", "PUNTER/KICKER"}) or st[1:] or st
    lss = _sub(st, {"LS", "LONG SNAPPER", "LONGSNAPPER", "DEEP SNAPPER", "SNAPPER"}) or st[2:] or st

    # returners: pin-only (no public data names returners in July) — pins are
    # matched against the WHOLE roster, any position
    all_players = [p for g in DEPTH_GROUPS for p in groups[g]]

    def _pin_pool(grp):
        pinned = overrides.get((norm(team), grp), [])
        by_norm = {norm(p["name"]): p for p in all_players}
        return [by_norm[n] for n in pinned if n in by_norm]

    rbs = groups["RB"]
    pools = {"QB": groups["QB"], "WR": groups["WR"], "TE": groups["TE"],
             "LT": p_lt, "LG": p_lg, "C": p_c, "RG": p_rg, "RT": p_rt,
             "DE": edges, "DT": inter, "NT": inter,
             "LB": lbs, "OLB": lbs[0::2], "ILB": lbs[1::2],
             "WLB": lbs[0::3], "MLB": lbs[1::3], "SLB": lbs[2::3],
             "CB": cbs, "NB": nbs, "S": safs, "SS": safs[0::2], "FS": safs[1::2],
             "K": ks, "P": ps, "LS": lss,
             "KR": _pin_pool("KR"), "PR": _pin_pool("PR")}
    if off_k == "21":
        pools["RB"], pools["FB"] = rbs[0::2], rbs[1::2]
    elif off_k == "flex":
        pools["FB"], pools["AB"] = rbs[0::3], [p for i, p in enumerate(rbs) if i % 3]
    else:
        pools["RB"] = rbs

    def _band(text):
        return ("", [text, "", "", ""])

    rows = ([_band(f"OFFENSE — {OFF_LABELS[off_k]} (projected)")]
            + _pad(_fill_rows(off, pools, all_pins), GRID_OFF)
            + [_band(f"DEFENSE — {DEF_LABELS[def_k]} (projected)")]
            + _pad(_fill_rows(deff, pools, all_pins), GRID_DEF)
            + [_band("SPECIAL TEAMS (projected)")]
            + _pad(_fill_rows(ST_ROWS, pools, all_pins), GRID_ST))
    label = f"{OFF_LABELS[off_k]} · {DEF_LABELS[def_k]} defense (projected)"
    return rows, label


def build_depth_rows(rosters: dict) -> dict[str, dict[str, str]]:
    """{team: {group: multiline depth text}} from roster + overrides."""
    overrides = load_depth_overrides()
    unmatched = []
    depth: dict[str, dict[str, str]] = {}
    for team, info in rosters.items():
        groups: dict[str, list] = {g: [] for g in DEPTH_GROUPS}
        for jersey, name, pos, cls, ht, wt, home, note in info["players"]:
            g = _group_of(pos)
            if not g or not name:
                continue
            groups[g].append(dict(name=name, pos=str(pos or "").strip(),
                                  cls=str(cls or "").strip(),
                                  jersey=jersey if isinstance(jersey, int) else 999,
                                  portal=str(note or "").startswith(("Transfer", "Incoming"))))
        team_out = {}
        for g, players in groups.items():
            pins = overrides.get((norm(team), g), [])

            def key(p):
                pn = norm(p["name"])
                pin = pins.index(pn) if pn in pins else 99
                return (pin, _cls_rank(p["cls"]), p["jersey"], p["name"])

            for pin_name in pins:
                if not any(norm(p["name"]) == pin_name for p in players):
                    unmatched.append(f"{team}/{g}: {pin_name}")
            players.sort(key=key)
            cap = DEPTH_CAPS[g]
            lines = []
            for i, p in enumerate(players[:cap], 1):
                pinned = "✓ " if norm(p["name"]) in pins else ""
                tag = " • portal" if p["portal"] else ""
                cls = f", {p['cls']}" if p["cls"] else ""
                lines.append(f"{i}. {pinned}{p['name']} ({p['pos']}{cls}){tag}")
            if len(players) > cap:
                lines.append(f"   +{len(players) - cap} more on roster")
            team_out[g] = "\n".join(lines) if lines else "(none listed)"
        depth[team] = team_out
    if unmatched:
        print(f"WARN depth overrides with no roster match: {unmatched[:8]}"
              + (f" (+{len(unmatched)-8})" if len(unmatched) > 8 else ""))
    return depth


def build_depth_sheet(wb, rosters: dict):
    """Hidden _Depth data sheet + visible 'Depth Charts' dropdown tab."""
    depth = build_depth_rows(rosters)

    if "_Depth" in wb.sheetnames:
        del wb["_Depth"]
    ws_d = wb.create_sheet("_Depth")
    ws_d.append(["Team"] + DEPTH_GROUPS)
    for team in sorted(depth):
        ws_d.append([team] + [depth[team][g] for g in DEPTH_GROUPS])
    ws_d.sheet_state = "hidden"

    prior = wb["Depth Charts"]["B1"].value if "Depth Charts" in wb.sheetnames else None
    if "Depth Charts" in wb.sheetnames:
        del wb["Depth Charts"]
    ws = wb.create_sheet("Depth Charts")
    teams = sorted(depth)
    ws["A1"] = "Select team:"
    ws["A1"].font = ARIAL_B
    ws["B1"] = prior if prior in teams else teams[0]
    ws["B1"].font = Font(name="Arial", bold=True, size=12, color="0000FF")
    ws["B1"].fill = PICK_FILL
    stamp = f"{datetime.date.today():%B %d, %Y}"
    ws["D1"] = (f"PROJECTED depth from official rosters (refreshed {stamp}): ordered by "
                "class seniority; ✓ = pinned via depth_overrides.csv. Teams publish real "
                "two-deeps in game weeks — pin corrections in the CSV and press REFRESH.")
    ws["D1"].font = Font(name="Arial", italic=True, size=9)
    for i, t in enumerate(teams, 1):
        ws.cell(row=i, column=27, value=t)  # AA hidden
    dv = DataValidation(type="list", formula1=f"=$AA$1:$AA${len(teams)}",
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws["B1"])
    ws["AB1"] = "=MATCH($B$1,_Depth!$A:$A,0)"
    ws["AB2"] = "=MATCH($B$1,_Teams!$A:$A,0)"
    for col in ("AA", "AB"):
        ws.column_dimensions[col].hidden = True

    ws.cell(row=3, column=1, value="=$B$1").font = TITLE_FONT
    for c in range(1, 3):
        ws.cell(row=3, column=c).fill = TITLE_FILL
    ws["A4"] = '=INDEX(_Teams!$B:$B,$AB$2)&"  ·  FPI "&INDEX(_Teams!$D:$D,$AB$2)&"  (#"&INDEX(_Teams!$C:$C,$AB$2)&")"'
    ws["A4"].font = Font(name="Arial", italic=True, color="5C6B7E")

    from openpyxl.styles import Alignment as _Al
    colors = {"QB": "2E6B4F", "RB": "2E6B4F", "WR": "2E6B4F", "TE": "2E6B4F",
              "OL": "2E6B4F", "DL/EDGE": "A3332C", "LB": "A3332C", "DB": "A3332C",
              "Specialists": "5C6B7E"}
    r = 6
    for gi, g in enumerate(DEPTH_GROUPS):
        col_letter = get_column_letter(2 + gi)  # _Depth col B..J
        c = ws.cell(row=r, column=1, value=g)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=colors[g])
        body = ws.cell(row=r + 1, column=1,
                       value=f'=INDEX(_Depth!${col_letter}:${col_letter},$AB$1)&""')
        body.font = ARIAL
        body.alignment = _Al(wrap_text=True, vertical="top")
        ws.row_dimensions[r + 1].height = 14.5 * (DEPTH_CAPS[g] + 1) + 6
        r += 3
    ws.column_dimensions["A"].width = 88
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A2"
    print(f"Depth Charts: {len(teams)} teams, 9 position groups")


# ---------------- scouting sheet ----------------

def build_scouting_sheet(wb):
    """'Scouting' tab: dropdown over the FPI top 25 -> full dossier card
    (base schemes, tendencies, strengths/weaknesses) from _Teams scout columns."""
    scouting = load_scouting()
    order = scouting["meta"].get("order", [])
    if not order:
        return
    prior = wb["Scouting"]["B1"].value if "Scouting" in wb.sheetnames else None
    if "Scouting" in wb.sheetnames:
        del wb["Scouting"]
    ws = wb.create_sheet("Scouting")

    ws["A1"] = "Select team:"
    ws["A1"].font = ARIAL_B
    ws["B1"] = prior if prior in order else order[0]
    ws["B1"].font = Font(name="Arial", bold=True, size=12, color="0000FF")
    ws["B1"].fill = PICK_FILL
    ws["D1"] = f"FPI top 25 + all SEC teams — {scouting['meta'].get('vintage', '')}"
    ws["D1"].font = Font(name="Arial", italic=True, size=9)
    for i, t in enumerate(order, 1):
        ws.cell(row=i, column=27, value=t)  # AA (hidden)
    dv = DataValidation(type="list", formula1=f"=$AA$1:$AA${len(order)}",
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws["B1"])
    ws["AB1"] = "=MATCH($B$1,_Teams!$A:$A,0)"
    for col in ("AA", "AB"):
        ws.column_dimensions[col].hidden = True

    ws.cell(row=3, column=1, value="=$B$1").font = TITLE_FONT
    for c in range(1, 3):
        ws.cell(row=3, column=c).fill = TITLE_FILL
    ws["A4"] = ('=INDEX(_Teams!$B:$B,$AB$1)&"  ·  FPI "&INDEX(_Teams!$D:$D,$AB$1)'
                '&"  (#"&INDEX(_Teams!$C:$C,$AB$1)&")"')
    ws["A4"].font = Font(name="Arial", italic=True, color="5C6B7E")

    sections = [
        ("OFFENSE — BASE", "AE", "2E6B4F", 48),
        ("OFFENSE — TENDENCIES", "AF", "2E6B4F", 72),
        ("DEFENSE — BASE", "AG", "A3332C", 48),
        ("DEFENSE — TENDENCIES", "AH", "A3332C", 72),
        ("STRENGTHS", "AI", "2E6B4F", 66),
        ("WEAKNESSES", "AJ", "A3332C", 66),
    ]
    from openpyxl.styles import Alignment as _Al
    r = 6
    for label, col, color, height in sections:
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=color)
        body = ws.cell(row=r + 1, column=1,
                       value=f'=INDEX(_Teams!${col}:${col},$AB$1)&""')
        body.font = ARIAL
        body.alignment = _Al(wrap_text=True, vertical="top")
        ws.row_dimensions[r + 1].height = height
        r += 3
    ws.column_dimensions["A"].width = 108
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A2"


# ---------------- upset board + alert log ----------------

ALERTS_LOG = HERE / "alerts_log.json"


def update_alerts_log(games: list[dict]) -> dict:
    """Persist each flagged game's FIRST-SEEN line so completed games are
    graded against the line we alerted on, not wherever it closed."""
    log = json.load(open(ALERTS_LOG, encoding="utf-8")) if ALERTS_LOG.exists() else {}
    today = f"{datetime.date.today():%Y-%m-%d}"
    for g in games:
        gid = str(g["id"])
        if g["tier"] and gid not in log:
            log[gid] = dict(first_seen=today, wk=g["wk"], home=g["home"],
                            away=g["away"], spread=g["spread"], edge=g["edge"],
                            tier=g["tier"], model_side=g["model_side"],
                            dog=g["dog"], dog_ml=g["dog_ml"],
                            ml_guard=g["ml_guard"])
        elif (gid in log and log[gid].get("tier") == "RED"
              and "dog_ml" not in log[gid] and not g["completed"]):
            # one-time backfill for RED entries logged before the ML guardrail
            # existed — allowed only while the game hasn't kicked off, so the
            # first-seen principle is preserved in spirit
            log[gid].update(dog=g["dog"], dog_ml=g["dog_ml"],
                            ml_guard=g["ml_guard"], ml_backfilled=today)
        # CLV tracking: keep the latest line each refresh. CFBD stops moving a
        # game's line at kickoff, so once the game completes this IS the close;
        # the first-seen number used for grading is never touched.
        if gid in log and g["spread"] is not None:
            log[gid]["close_spread"] = g["spread"]
    json.dump(log, open(ALERTS_LOG, "w", encoding="utf-8"), indent=0)
    return log


def _grade(entry: dict, g: dict) -> tuple[str, str]:
    """(ats, outright) for the logged model side vs the logged spread."""
    if not g["completed"] or g["home_pts"] is None:
        return "", ""
    margin = g["home_pts"] - g["away_pts"]  # home perspective
    covered = margin + float(entry["spread"])  # >0 = home covered
    side_is_home = entry["model_side"] == entry["home"]
    ats_val = covered if side_is_home else -covered
    ats = "Push" if ats_val == 0 else ("Cover" if ats_val > 0 else "Miss")
    won = margin > 0 if side_is_home else margin < 0
    outright = "Dog WON" if entry["tier"] == "RED" and won else \
               ("Dog lost" if entry["tier"] == "RED" else "")
    return ats, outright


def build_upset_board(wb, games: list[dict]):
    log = update_alerts_log(games)
    by_id = {str(g["id"]): g for g in games}

    if "Upset Board" in wb.sheetnames:
        del wb["Upset Board"]
    ws = wb.create_sheet("Upset Board")

    ws["A1"] = "Upset Board — FPI vs the market, 2026"
    ws["A1"].font = TITLE_FONT
    for c in range(1, 15):
        ws.cell(row=1, column=c).fill = TITLE_FILL
    ws["A2"] = ("Model margin = FPI gap + 2.5 home field. PRIOR = ESPN 2026 PRESEASON FPI (captured from ESPN "
                "July 14, 2026). \U0001F534 model picks the underdog outright (spread >= 3). "
                "\U0001F7E1 same side, 6+ pt disagreement. Games are graded against the line first seen at alert time. "
                "2023-25 backtest of these rules (with a stale prior): 49.7% ATS — this board is a research "
                "shortlist and narrative engine, not a bet slip. See BACKTEST_RESULTS.md.")
    ws["A2"].font = Font(name="Arial", italic=True, size=9)
    tail_thresh = next((g["ou_tail_thresh"] for g in games
                        if g.get("ou_tail_thresh") is not None), None)
    ws["A3"] = ("Post-mortem rules (market-postmortem/MARKET_POSTMORTEM.md): ⚑U-TAIL = total in the season's "
                f"top decile{f' (>= {tail_thresh:g})' if tail_thresh else ''} — unders went 55.1% on 60+ totals "
                "2021-25, the only spread/total bias to survive FDR. ⛔ATS-only = RED dog at +401 or longer on the "
                "road/neutral — longshot moneylines bled -22.9% ROI 2021-25; take the points, not the ML. "
                "Home dogs exempt (-1.2%). CLV = points the alert line beat the latest/closing line by, from the "
                "model side's perspective (2021-25: no strategy beat the close, so positive CLV is the cleanest "
                "sign the alert engine is early rather than loud).")
    ws["A3"].font = Font(name="Arial", italic=True, size=9)

    headers = ["Wk", "Date", "Matchup", "Spread (alert)", "Now", "CLV",
               "O/U", "FPI margin (home)", "Edge", "Tier", "Dog ML",
               "Model side", "Final", "ATS"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL

    RED_F = PatternFill("solid", start_color="FFC7CE")
    YEL_F = PatternFill("solid", start_color="FFEB9C")

    rows = []
    for gid, e in log.items():
        g = by_id.get(gid)
        if g is None:
            continue
        ats, outright = _grade(e, g)
        rows.append((e, g, ats, outright))
    rows.sort(key=lambda t: (t[0]["wk"] or 0, -abs(t[0]["edge"] or 0)))

    r = 5
    ats_rec = {"Cover": 0, "Miss": 0, "Push": 0}
    dog_rec = {"Dog WON": 0, "Dog lost": 0}
    clv_done: list[float] = []
    for e, g, ats, outright in rows:
        matchup = f"{e['away']} at {e['home']}" if not g["neutral"] else \
                  f"{e['away']} vs {e['home']} (N)"
        final = _result_text(g, True) and f"{g['home_pts']}-{g['away_pts']}"
        if e.get("ml_guard"):
            outright = ""  # guardrail: the alert is ATS-only, no ML framing
        ou_disp = ""
        if g["ou"] is not None:
            ou_disp = (f"{float(g['ou']):g} ⚑U-TAIL" if g.get("ou_tail")
                       else g["ou"])
        dog_ml_disp = ""
        if e["tier"] == "RED":
            dog_ml_disp = _ml_text(e.get("dog_ml")) or "n/a"
            if e.get("ml_guard"):
                dog_ml_disp += " ⛔ATS-only"
        clv = ""
        if e.get("close_spread") is not None and e.get("spread") is not None:
            moved = float(e["spread"]) - float(e["close_spread"])
            clv = round(moved if e["model_side"] == e["home"] else -moved, 1)
        vals = [e["wk"], g["date"], matchup,
                _spread_text(e["home"], e["away"], e["spread"]),
                g["spread_text"], clv, ou_disp,
                round(g["model_margin"], 1) if g["model_margin"] is not None else "",
                e["edge"], "\U0001F534" if e["tier"] == "RED" else "\U0001F7E1",
                dog_ml_disp, e["model_side"], final,
                (ats + (" / " + outright if outright else "")) if ats else "pending"]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = ARIAL
            c.fill = RED_F if e["tier"] == "RED" else YEL_F
        if ats:
            ats_rec[ats] += 1
        if outright:
            dog_rec[outright] += 1
        if clv != "" and g["completed"]:
            clv_done.append(clv)
        r += 1

    # scorecard
    ws.cell(row=4, column=16, value="Scorecard").font = ARIAL_B
    n_guard = sum(1 for e, *_ in rows if e.get("ml_guard"))
    if clv_done:
        clv_avg = f"{sum(clv_done) / len(clv_done):+.2f}"
        clv_rec = (f"{sum(1 for c in clv_done if c > 0)}-"
                   f"{sum(1 for c in clv_done if c < 0)}-"
                   f"{sum(1 for c in clv_done if c == 0)}")
    else:
        clv_avg, clv_rec = "pending", "0-0-0"
    score_lines = [
        ("Alerts logged", len(rows)),
        ("Model side ATS", f"{ats_rec['Cover']}-{ats_rec['Miss']}-{ats_rec['Push']}"),
        ("Red dogs outright*", f"{dog_rec['Dog WON']}-{dog_rec['Dog lost']}"),
        ("ML-guarded (ATS-only)", n_guard),
        ("Avg CLV, graded (pts)", clv_avg),
        ("CLV beat-tie-lost close", clv_rec),
        ("Pending", sum(1 for *_x, a, _o in rows if not a)),
        ("*excludes ⛔ML-guarded", ""),
    ]
    for i, (k, v) in enumerate(score_lines):
        ws.cell(row=5 + i, column=16, value=k).font = ARIAL
        ws.cell(row=5 + i, column=17, value=v).font = ARIAL_B

    widths = [5, 24, 34, 16, 16, 6, 14, 16, 8, 6, 16, 20, 10, 16, 2, 22, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    print(f"Upset Board: {len(rows)} alerts "
          f"({sum(1 for e, *_ in rows if e['tier'] == 'RED')} red)")


# ---------------- watch list (paper bets) ----------------

WATCH_LOG = HERE / "watchlist_log.json"
POWER_CONFS_2026 = {"SEC", "Big Ten", "Big 12", "ACC"}

WATCH_RULES = {
    "RANKED_FAV": ("Ranked-vs-ranked favorite ATS",
                   "2021-25: 56.1% (285 gms), 5/5 seasons, p=.048 — failed FDR"),
    "EARLY_UNDER": ("Early-season under (wks 1-4)",
                    "2021-25: 52.8% (1,218 bets), 5/5 seasons — at break-even"),
    "G5_DOG": ("G5-vs-G5 dog ATS",
               "2021-25: 52.1% (1,622 bets), 4/5 seasons — short of break-even"),
}


def fetch_ap_ranks(refresh: bool) -> dict[int, dict[str, int]]:
    """{week: {norm(team): rank}} from the 2026 AP poll (Coaches fallback).

    Empty until the first 2026 poll drops (~late Aug) — RANKED_FAV simply
    logs nothing before then.
    """
    import cfbd_client as cfbd
    try:
        weeks = cfbd.get("/rankings", {"year": 2026, "seasonType": "regular"},
                         refresh)
    except Exception as exc:
        print(f"watch list: 2026 rankings unavailable ({exc})")
        return {}
    out = {}
    for wk in weeks or []:
        polls = wk.get("polls") or []
        poll = (next((p for p in polls if p.get("poll") == "AP Top 25"), None)
                or next((p for p in polls if p.get("poll") == "Coaches Poll"),
                        None))
        if poll:
            out[wk.get("week")] = {norm(r["school"]): r.get("rank")
                                   for r in poll.get("ranks") or []
                                   if r.get("school")}
    return out


def _is_g5(conf, cls, team) -> bool:
    return (cls == "fbs" and conf not in POWER_CONFS_2026
            and norm(team or "") != norm("Notre Dame"))


def update_watch_log(games: list[dict], ranks: dict) -> dict:
    """Paper-bet ledger for the post-mortem watch-list patterns: log each rule
    hit at the FIRST-SEEN line, pre-kickoff only, never regraded against moved
    lines (same philosophy as alerts_log). A game can hit multiple rules."""
    log = json.load(open(WATCH_LOG, encoding="utf-8")) if WATCH_LOG.exists() else {}
    today = f"{datetime.date.today():%Y-%m-%d}"

    def put(g, rule, **kw):
        key = f"{g['id']}:{rule}"
        if key not in log:
            log[key] = dict(first_seen=today, rule=rule, gid=g["id"],
                            wk=g["wk"], home=g["home"], away=g["away"], **kw)

    for g in games:
        if g["completed"]:
            continue  # paper bets must exist before kickoff
        fbs_both = g["home_class"] == "fbs" and g["away_class"] == "fbs"
        spread, ou = g["spread"], g["ou"]
        poll = ranks.get(g["wk"]) or {}
        if (fbs_both and spread is not None and spread != 0 and poll
                and norm(g["home"]) in poll and norm(g["away"]) in poll):
            fav = g["home"] if spread < 0 else g["away"]
            put(g, "RANKED_FAV", side=fav, spread=spread,
                bet=f"{_spread_text(g['home'], g['away'], spread)} ATS")
        if fbs_both and ou is not None and (g["wk"] or 99) <= 4:
            put(g, "EARLY_UNDER", side="UNDER", total=ou,
                bet=f"UNDER {float(ou):g}")
        if (spread is not None and spread != 0
                and _is_g5(g["home_conf"], g["home_class"], g["home"])
                and _is_g5(g["away_conf"], g["away_class"], g["away"])):
            dog = g["away"] if spread < 0 else g["home"]
            put(g, "G5_DOG", side=dog, spread=spread,
                bet=f"{dog} +{abs(float(spread)):g} ATS")
    json.dump(log, open(WATCH_LOG, "w", encoding="utf-8"), indent=0)
    return log


def _grade_watch(e: dict, g: dict) -> str:
    """W/L/P vs the first-seen number; '' while pending."""
    if not g["completed"] or g["home_pts"] is None:
        return ""
    if e["rule"] == "EARLY_UNDER":
        diff = float(e["total"]) - (g["home_pts"] + g["away_pts"])
        return "P" if diff == 0 else ("W" if diff > 0 else "L")
    margin = g["home_pts"] - g["away_pts"]
    covered = margin + float(e["spread"])  # >0 = home covered
    val = covered if e["side"] == e["home"] else -covered
    return "P" if val == 0 else ("W" if val > 0 else "L")


def build_watch_list(wb, games: list[dict], refresh: bool):
    ranks = fetch_ap_ranks(refresh)
    log = update_watch_log(games, ranks)
    by_id = {g["id"]: g for g in games}

    if "Watch List" in wb.sheetnames:
        del wb["Watch List"]
    ws = wb.create_sheet("Watch List")
    ws["A1"] = "Watch List — post-mortem patterns on paper, 2026"
    ws["A1"].font = TITLE_FONT
    for c in range(1, 9):
        ws.cell(row=1, column=c).fill = TITLE_FILL
    ws["A2"] = ("Three patterns from the 2021-25 market post-mortem that repeated across seasons but did NOT "
                "survive multiple-comparison correction (market-postmortem/MARKET_POSTMORTEM.md §4). Logged as "
                "PAPER bets at the first-seen line, pre-kickoff only, graded like the Upset Board. Graduation "
                "rule: a pattern earns real consideration only if it clears 52.38% on 100+ decided 2026 bets; "
                "otherwise it retires as 2021-25 noise. Evidence collection, not a bet slip.")
    ws["A2"].font = Font(name="Arial", italic=True, size=9)

    rows = []
    for key, e in log.items():
        g = by_id.get(e["gid"])
        if g is not None:
            rows.append((e, g, _grade_watch(e, g)))
    rule_order = {r: i for i, r in enumerate(WATCH_RULES)}
    rows.sort(key=lambda t: (rule_order.get(t[0]["rule"], 9), t[0]["wk"] or 0))

    # per-pattern scorecard (cols A, D-G; D is wide enough for the evidence)
    for j, h in ((1, "Pattern"), (4, "2021-25 evidence"), (5, "2026 paper record"),
                 (6, "Win%"), (7, "Status")):
        c = ws.cell(row=4, column=j, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
    for i, (rule, (label, hist)) in enumerate(WATCH_RULES.items()):
        graded = [gr for e, _g, gr in rows if e["rule"] == rule]
        w, l, p = graded.count("W"), graded.count("L"), graded.count("P")
        n = w + l
        pct = 100 * w / n if n else None
        if n >= 100 and pct > 52.38:
            status = "GRADUATING — clears break-even"
        elif n >= 100 and pct <= 50:
            status = "RETIRING — below coin flip"
        elif len(graded) or any(e["rule"] == rule for e, *_ in rows):
            status = "monitoring"
        else:
            status = "waiting for games"
        for j, v in ((1, label), (4, hist), (5, f"{w}-{l}-{p}"),
                     (6, f"{pct:.1f}%" if pct is not None else "—"),
                     (7, status)):
            ws.cell(row=5 + i, column=j, value=v).font = ARIAL

    hrow = 9
    for j, h in enumerate(["Pattern", "Wk", "Date", "Matchup",
                           "Paper bet (first-seen)", "Now", "Final",
                           "Result"], 1):
        c = ws.cell(row=hrow, column=j, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
    r = hrow + 1
    for e, g, gr in rows:
        matchup = f"{e['away']} at {e['home']}" if not g["neutral"] else \
                  f"{e['away']} vs {e['home']} (N)"
        now = ((f"{float(g['ou']):g}" if g["ou"] is not None else "")
               if e["rule"] == "EARLY_UNDER" else g["spread_text"])
        final = (f"{g['home_pts']}-{g['away_pts']}"
                 if g["completed"] and g["home_pts"] is not None else "")
        vals = [WATCH_RULES[e["rule"]][0], e["wk"], g["date"], matchup,
                e["bet"], now, final,
                {"W": "Win", "L": "Loss", "P": "Push"}.get(gr, "pending")]
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j, value=v).font = ARIAL
        r += 1

    widths = [30, 5, 24, 42, 26, 16, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hrow + 1}"
    counts = {r_: sum(1 for e, *_ in rows if e["rule"] == r_)
              for r_ in WATCH_RULES}
    print(f"Watch List: {len(rows)} paper bets {counts}")


# ---------------- season sim ----------------

def build_season_sim(wb, games: list[dict], fpi: dict[str, dict],
                     team_conf: dict[str, str], n_sims: int = 10000):
    import numpy as np

    from statistics import NormalDist

    SIGMA = 13.5  # historical sd of CFB scoring margin vs spread
    ND = NormalDist(0, SIGMA)
    teams = sorted(team_conf)
    win_probs: dict[str, list[float]] = {t: [] for t in teams}
    for g in games:
        for team, opp, is_home in ((g["home"], g["away"], True),
                                   (g["away"], g["home"], False)):
            if team not in win_probs:
                continue
            tf, of = fpi.get(norm(team)), fpi.get(norm(opp))
            if tf is None:
                continue
            if of is None:
                margin = UNRATED_MARGIN
            else:
                margin = tf["fpi"] - of["fpi"]
            if not g["neutral"]:
                margin += HFA if is_home else -HFA
            win_probs[team].append(ND.cdf(margin))

    rng = np.random.default_rng(2026)
    results = []
    for t in teams:
        ps = win_probs[t]
        if not ps:
            continue
        sims = (rng.random((n_sims, len(ps))) < np.array(ps)).sum(axis=1)
        results.append(dict(
            team=t, conf=team_conf[t],
            fpi=fpi.get(norm(t), {}).get("fpi", "n/a"),
            games=len(ps), mean=float(sims.mean()),
            p10=int(np.percentile(sims, 10)), p90=int(np.percentile(sims, 90)),
            bowl=float((sims >= 6).mean()), ten=float((sims >= 10).mean())))
    results.sort(key=lambda d: -d["mean"])

    if "Season Sim" in wb.sheetnames:
        del wb["Season Sim"]
    ws = wb.create_sheet("Season Sim")
    ws["A1"] = "2026 Season Simulation — 10,000 Monte Carlo runs"
    ws["A1"].font = TITLE_FONT
    for c in range(1, 10):
        ws.cell(row=1, column=c).fill = TITLE_FILL
    ws["A2"] = ("Win prob per game = NormalDist(FPI gap + 2.5 HFA, sd 13.5). Prior = ESPN 2026 PRESEASON FPI "
                "(captured July 14, 2026). Unrated opponents counted as +24 pt margin. "
                "Compare Proj Wins to your book's season win totals for value.")
    ws["A2"].font = Font(name="Arial", italic=True, size=9)
    headers = ["Rank", "Team", "Conf", "FPI", "Games", "Proj Wins",
               "10th pct", "90th pct", "P(6+ / bowl)", "P(10+)"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
    for i, d in enumerate(results, 1):
        vals = [i, d["team"], d["conf"], d["fpi"], d["games"],
                round(d["mean"], 1), d["p10"], d["p90"],
                d["bowl"], d["ten"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=4 + i, column=j, value=v)
            c.font = ARIAL
            if j in (9, 10):
                c.number_format = "0%"
    for i, w in enumerate([6, 24, 18, 8, 7, 10, 9, 9, 12, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:J{4 + len(results)}"
    print(f"Season Sim: {len(results)} teams projected")


def restructure(book: Path, refresh: bool = False, drop_team_tabs: bool = True):
    keep_vba = book.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(book, keep_vba=keep_vba)
    stamp = f"{datetime.date.today():%B %d, %Y}"

    # remember current dropdown picks
    prior = {}
    for conf in CONF_ORDER:
        if conf in wb.sheetnames:
            prior[conf] = wb[conf]["B1"].value

    conf_teams, games, fpi, team_conf = build_data_sheets(wb, refresh)
    build_upset_board(wb, games)
    build_watch_list(wb, games, refresh)
    build_season_sim(wb, games, fpi, team_conf)
    # standalone Scouting / Depth Charts tabs retired 2026-07-13: their content
    # now lives on the conference tabs (S&W + scheme depth panel per team)

    # per-conference max roster size drives each tab's formula-grid height
    ws_t = wb["_Teams"]
    counts: dict[str, int] = {}
    for row in ws_t.iter_rows(min_row=2, values_only=True):
        counts[row[1]] = max(counts.get(row[1], 0), row[27] or 0)

    for conf in CONF_ORDER:
        teams = sorted(conf_teams.get(conf, []))
        if not teams:
            continue
        build_conference_tab(wb, conf, teams, counts.get(conf, 130), stamp,
                             prior.get(conf))

    # order: Overview, FPI Decomposition, conferences, hidden data sheets
    if drop_team_tabs:
        keep = {"Overview", "FPI Decomposition", "Upset Board", "Watch List",
                "Season Sim", "_Teams", "_Rosters", "_Sched", "_DepthGrid"}
        keep |= set(CONF_ORDER)
        for name in list(wb.sheetnames):
            if name not in keep:
                del wb[name]
    order = ["Overview", "FPI Decomposition", "Upset Board", "Watch List",
             "Season Sim"] + \
            [c for c in CONF_ORDER if c in wb.sheetnames] + \
            ["_Teams", "_Rosters", "_Sched", "_DepthGrid"]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames] + \
                 [wb[n] for n in wb.sheetnames if n not in order]
    wb.save(book)
    n_teams = sum(len(v) for v in conf_teams.values())
    print(f"restructured: {len([c for c in CONF_ORDER if c in wb.sheetnames])} "
          f"conference tabs covering {n_teams} teams")


if __name__ == "__main__":
    from refresh_all import team_book, load_env_key
    load_env_key()
    restructure(team_book(), refresh="--refresh" in sys.argv)
