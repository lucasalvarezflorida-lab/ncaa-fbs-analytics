"""Build NFL_Postmortem.xlsx and NBA_Postmortem.xlsx from the phase-2 outputs.

Each workbook: Summary (live COUNTIFS/AVERAGE/SLOPE formulas over the Bets
sheet), Slice Results (the full test battery — p-values/CIs/BH flags are
static analysis outputs from analyze_nfl.py / analyze_nba.py), Bets (the
per-game dataset with autofilter), Charts (embedded PNGs).

After running, recalc via Excel COM (NOT LibreOffice — on this machine it
writes empty cached values; see cfb-pipeline skill).
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent

ARIAL = Font(name="Arial", size=10)
ARIAL_B = Font(name="Arial", size=10, bold=True)
ARIAL_I = Font(name="Arial", size=9, italic=True, color="52514E")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", start_color="1F3864")
HEAD_FILL = PatternFill("solid", start_color="2F5496")
WHITE_B = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SIG_FILL = PatternFill("solid", start_color="C6EFCE")   # BH-significant
PERSIST_FILL = PatternFill("solid", start_color="FFEB9C")  # persistent only
PCT = "0.0%"
NUM1 = "0.0"


def title_block(ws, title: str, notes: list[str], width: int) -> int:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    for c in range(1, width + 1):
        ws.cell(row=1, column=c).fill = TITLE_FILL
    r = 2
    for note in notes:
        ws.cell(row=r, column=1, value=note).font = ARIAL_I
        r += 1
    return r + 1


def write_table(ws, df: pd.DataFrame, start_row: int, sig_col: str | None = None,
                persist_col: str | None = None) -> None:
    for j, h in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        fill = None
        d = dict(zip(df.columns, row))
        if sig_col and d.get(sig_col) == 1:
            fill = SIG_FILL
        elif persist_col and d.get(persist_col) == 1:
            fill = PERSIST_FILL
        for j, v in enumerate(row, 1):
            if isinstance(v, float) and np.isnan(v):
                v = None
            c = ws.cell(row=i, column=j, value=v)
            c.font = ARIAL
            if fill:
                c.fill = fill
    ws.auto_filter.ref = (f"A{start_row}:"
                          f"{get_column_letter(len(df.columns))}"
                          f"{start_row + len(df)}")
    ws.freeze_panes = f"A{start_row + 1}"


def summary_sheet(ws, sport: dict, bets: pd.DataFrame, col: dict) -> None:
    n = len(bets)
    last = n + 1  # data ends on this sheet row

    def rng(name: str) -> str:
        return f"Bets!${col[name]}$2:${col[name]}${last}"

    r = title_block(ws, sport["title"], sport["notes"], 8)

    def section(label: str) -> None:
        nonlocal r
        c = ws.cell(row=r, column=1, value=label)
        c.font = WHITE_B
        c.fill = HEAD_FILL
        for j in range(2, 9):
            ws.cell(row=r, column=j).fill = HEAD_FILL
        r += 1

    def line(label: str, value, fmt: str | None = None, note: str = "") -> None:
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = ARIAL
        c = ws.cell(row=r, column=3, value=value)
        c.font = ARIAL_B
        if fmt:
            c.number_format = fmt
        if note:
            ws.cell(row=r, column=4, value=note).font = ARIAL_I
        r += 1

    section("Dataset")
    line("Games graded", f"=COUNTA(Bets!$A$2:$A${last})")
    line("Window", sport["window"])
    r += 1

    section("Calibration — live formulas over the Bets tab")
    line("Spread bias, actual - line (pts)", f"=AVERAGE({rng('home_cover_margin')})",
         "0.000", "positive = home sides beat the closing spread")
    line("Spread MAE (pts)", f"=AVERAGE({rng('abs_spread_err')})", "0.00")
    line("Spread slope (fair = 1.00)",
         f"=SLOPE({rng('margin')},{rng('pred_home_margin')})", "0.000")
    line("Total bias, actual - line (pts)", f"=AVERAGE({rng('over_margin')})",
         "0.000")
    line("Total MAE (pts)", f"=AVERAGE({rng('abs_total_err')})", "0.00")
    line("Total slope (fair = 1.00)",
         f"=SLOPE({rng('total_points')},{rng('total_close')})", "0.000")
    r += 1

    section("Report card vs the close — break-even at -110 = 52.38%")
    hdr = ws.cell(row=r, column=3, value="W")
    for j, h in ((3, "W"), (4, "L"), (5, "P"), (6, "win%")):
        c = ws.cell(row=r, column=j, value=h)
        c.font = ARIAL_B
    r += 1

    def record(label: str, res_rng: str, extra: str = "",
               outcomes: tuple = ("W", "L", "P")) -> None:
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = ARIAL
        for j, out in zip((3, 4, 5), outcomes):
            f = f'=COUNTIFS({res_rng},"{out}"{extra})'
            ws.cell(row=r, column=j, value=f).font = ARIAL
        wc, lc = f"C{r}", f"D{r}"
        c = ws.cell(row=r, column=6, value=f"={wc}/({wc}+{lc})")
        c.font = ARIAL_B
        c.number_format = "0.00%"
        r += 1

    neutral_extra = (f',Bets!${col["neutral"]}$2:${col["neutral"]}${last},0'
                     if "neutral" in col else "")
    record("Home ATS" + (" (non-neutral)" if "neutral" in col else ""),
           rng("home_covered"), neutral_extra)
    record("Dog ATS", rng("dog_res"))
    ws.cell(row=r - 1, column=7,
            value="(fav ATS is the mirror image)").font = ARIAL_I
    record("Under (W = under cashed)", rng("over_result"),
           outcomes=("U", "O", "P"))
    r += 1
    line("Favorites ML, flat-bet ROI per $1", f"=AVERAGE({rng('ret_fav')})",
         PCT, "closing prices; blank rows (no ML) excluded")
    line("Dogs ML, flat-bet ROI per $1", f"=AVERAGE({rng('ret_dog')})", PCT)
    r += 1

    section("What survived Benjamini-Hochberg (q=0.10)")
    for txt in sport["survivors"]:
        ws.cell(row=r, column=1, value=txt).font = ARIAL
        r += 1
    r += 1
    section("Watch list — persistent, NOT proven (fails FDR)")
    for txt in sport["watchlist"]:
        ws.cell(row=r, column=1, value=txt).font = ARIAL
        r += 1
    r += 1
    section("Caveats")
    for txt in sport["caveats"]:
        ws.cell(row=r, column=1, value=txt).font = ARIAL
        r += 1

    for i, w in enumerate([44, 2, 10, 10, 8, 9, 40, 6], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_book(sport: dict) -> Path:
    bets = pd.read_csv(HERE / sport["bets_csv"])
    bets = bets[sport["bet_cols"]].copy()
    slices = pd.read_csv(HERE / sport["slices_csv"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws_b = wb.create_sheet("Bets")
    col = {name: get_column_letter(j)
           for j, name in enumerate(sport["bet_cols"], 1)}
    write_table(ws_b, bets, 1)
    for name in ("spread_close", "spread_open", "total_close", "total_open",
                 "home_cover_margin", "over_margin", "abs_spread_err",
                 "abs_total_err", "pred_home_margin", "spread_move"):
        if name in col:
            for c in ws_b[col[name]][1:]:
                c.number_format = NUM1
    for name in ("ret_fav", "ret_dog"):
        for c in ws_b[col[name]][1:]:
            c.number_format = "0.00"
    ws_b.column_dimensions[col["date"]].width = 12
    for name in ("home", "away"):
        ws_b.column_dimensions[col[name]].width = 16

    ws_s = wb.create_sheet("Slice Results")
    r0 = title_block(
        ws_s, f"{sport['name']} — all {len(slices)} strategy tests",
        ["Static outputs of analyze_" + sport["key"] + ".py: exact binomial p vs 50%, Wilson 95% CI (ATS/totals) or bootstrap "
         "ROI CI (moneylines), Benjamini-Hochberg FDR q=0.10 across the whole battery.",
         "GREEN fill = survives BH correction. YELLOW = persistent across seasons but fails FDR (watch list). "
         "Everything else is noise. roi_at_110 = % per $1 staked.",
         "Source: " + sport["slices_csv"] + " (regenerate via the pipeline in README.md)."],
        min(len(slices.columns), 12))
    write_table(ws_s, slices, r0, sig_col="bh_significant",
                persist_col="persistent")

    ws_c = wb.create_sheet("Charts")
    ws_c["A1"] = f"{sport['name']} charts (from charts/, regenerable)"
    ws_c["A1"].font = ARIAL_B
    anchor_row = 3
    for png in sport["charts"]:
        img = XLImage(str(HERE / "charts" / png))
        img.width, img.height = int(img.width * 0.55), int(img.height * 0.55)
        ws_c.add_image(img, f"A{anchor_row}")
        anchor_row += 26

    summary_sheet(ws, sport, bets, col)

    out = HERE / sport["out"]
    wb.save(out)
    print(f"wrote {out.name}: {len(bets)} bets, {len(slices)} tests")
    return out


def prep_common(df: pd.DataFrame) -> pd.DataFrame:
    df["pred_home_margin"] = -df.spread_close
    df["abs_spread_err"] = df.home_cover_margin.abs()
    df["abs_total_err"] = df.over_margin.abs()
    return df


NFL_COLS = ["game_id", "season", "game_type", "week", "date", "weekday",
            "primetime", "home", "away", "neutral", "div_game", "roof",
            "temp", "wind", "home_rest", "away_rest", "spread_close",
            "pred_home_margin", "total_close", "home_ml", "away_ml", "margin",
            "total_points", "home_cover_margin", "abs_spread_err",
            "home_covered", "over_margin", "abs_total_err", "over_result",
            "dog_res", "abs_spread", "ml_fav", "ml_dog", "ret_fav", "ret_dog"]

NBA_COLS = ["season", "date", "home", "away", "home_pts", "away_pts",
            "home_rest", "away_rest", "spread_open", "spread_close",
            "pred_home_margin", "spread_move", "total_open", "total_close",
            "total_q", "home_ml", "away_ml", "margin", "total_points",
            "home_cover_margin", "abs_spread_err", "home_covered",
            "over_margin", "abs_total_err", "over_result", "dog_res",
            "dog_res_open", "abs_spread", "ml_fav", "ml_dog",
            "ret_fav", "ret_dog"]


def main() -> None:
    for csv, cols in (("nfl_bets_2021_2025.csv", None),
                      ("nba_bets_2011_2021.csv", None)):
        df = pd.read_csv(HERE / csv)
        prep_common(df).to_csv(HERE / csv, index=False)

    nfl = dict(
        key="nfl", name="NFL market post-mortem 2021-2025",
        title="NFL Market Post-Mortem — closing lines vs results, 2021-2025",
        window="2021-2025 incl. playoffs (nflverse closing lines)",
        bets_csv="nfl_bets_2021_2025.csv", slices_csv="nfl_slice_results.csv",
        bet_cols=NFL_COLS, out="NFL_Postmortem.xlsx",
        charts=["nfl_calibration_curves.png", "phase2_favorite_longshot.png",
                "phase2_totals_tails.png"],
        notes=["Source: nfl_games.csv (nflverse) -> analyze_nfl.py. Spread is home-perspective: negative = home favored.",
               "Full method + findings: MARKET_POSTMORTEM_PHASE2.md. Summary numbers below are LIVE formulas over the Bets tab.",
               "Verdict: 0 of 42 slices survive FDR correction — the sharpest board tested."],
        survivors=["(none — every famous angle was priced: byes, division dogs, primetime sides, weather, every spread bucket, every ML band)"],
        watchlist=["Primetime unders: 162-129-4 (55.7%), 5/5 seasons, p=.060",
                   "Dogs getting 7-9.5: 123-97-2 (55.9%) ATS, 4/5 seasons",
                   "Weeks 1-4 unders: 175-144-1 (54.9%), 5/5 seasons",
                   "Wind 15+ mph unders: 57-36-2 (61.3%) — 93 games, anecdote-grade"],
        caveats=["-110 assumed both ways (actual posted juice averages slightly better).",
                 "Home teams beat the close by +0.56 pts/game (1.7 sigma - suggestive, unproven).",
                 "One book per game, no line shopping. Neutral-site games excluded from Home ATS."])

    nba = dict(
        key="nba", name="NBA market post-mortem 2011-2021",
        title="NBA Market Post-Mortem — closing lines vs results, 2011-12 to 2021-22",
        window="Seasons 2011-2021 incl. playoffs (local odds archive)",
        bets_csv="nba_bets_2011_2021.csv", slices_csv="nba_slice_results.csv",
        bet_cols=NBA_COLS, out="NBA_Postmortem.xlsx",
        charts=["nba_calibration_curves.png", "phase2_favorite_longshot.png",
                "phase2_totals_tails.png"],
        notes=["Source: nba_archive.json -> analyze_nba.py. Spread is home-perspective: negative = home favored.",
               "933 games had close spread/total swapped in-source: totals repaired (validated), spread signs unrecoverable -> blank.",
               "Full method + findings: MARKET_POSTMORTEM_PHASE2.md. Summary numbers below are LIVE formulas over the Bets tab.",
               "total_q = within-season closing-total quintile (0 lowest .. 4 highest) — use it, not absolute totals (30-pt era drift)."],
        survivors=["Every favorites-ML band: ROI -4.0% to -5.1%, 11/11 seasons — this is the VIG measured precisely, not a bias (dogs lost the same or less; no CFB-style tail).",
                   "Follow totals steam (>=1 pt move): 51.4% vs the close, 9/11 seasons — informative, still sub-break-even.",
                   "Dogs vs the OPENING line: 48.9% (favorites gained value open->close) — NBA openers shade toward dogs, mirror of CFB."],
        watchlist=["Unders on each season's top total quintile: 51.6%, 10/11 seasons, p=.115 — faint echo of the CFB tail bias"],
        caveats=["Window predates the legal-betting liquidity era; today's board is likely sharper.",
                 "No playoff flag in the archive (month splits stand in). Lockout 2011 and COVID 2019/2020 seasons left in.",
                 "No spread juice in the archive: -110 assumed. One book per game."])

    for sport in (nfl, nba):
        build_book(sport)


if __name__ == "__main__":
    main()
