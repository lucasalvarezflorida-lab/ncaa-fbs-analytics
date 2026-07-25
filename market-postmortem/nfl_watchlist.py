"""2026 NFL watch-list tracker — the paper-bet loop for the four rules
frozen in NFL_WATCHLIST_2026_PREREG.md (run it any time; weekly in season).

What it does, per the prereg:
  * downloads the current nflverse games.csv (the live upstream of the
    static nfl_games.csv snapshot the backtest used),
  * logs PROVISIONAL entries for upcoming 2026 games whose posted line
    qualifies (first-seen line kept for information/CLV only),
  * grades completed games against the CLOSING line — eligibility AND
    result both use the close, exactly like the 2021-25 backtest
    (analyze_nfl.py definitions, sign conventions included). A provisional
    entry whose line drifts out of its bucket by close is marked dropped,
    not graded, and kept for audit.
  * maintains market-postmortem/nfl_watchlist_log.json and rebuilds the
    "Watch List 2026" tab in NFL_Postmortem.xlsx (values only — rerun this
    script after build_phase2_workbooks.py, which recreates the book).

Graduation (frozen): 52.38% on 100+ decided bets accumulated from 2026
week 1; no interim graduations, no mid-season rule edits.
"""
import datetime
import json
from pathlib import Path

import openpyxl
import pandas as pd
import requests
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
GAMES_URL = "https://github.com/nflverse/nfldata/raw/master/data/games.csv"
LIVE_CSV = HERE / "nfl_games_live.csv"
LOG = HERE / "nfl_watchlist_log.json"
BOOK = HERE / "NFL_Postmortem.xlsx"
SEASON = 2026

RULES = {
    "PRIME_UNDER": ("Under — primetime (Thu/Mon/SNF)",
                    "2021-25: 162-129-4, 55.7%, p=.060, 5/5 seasons"),
    "DOG_7_95": ("Dog ATS — getting 7-9.5",
                 "2021-25: 123-97-2, 55.9%, p=.092, 4/5 seasons"),
    "EARLY_UNDER": ("Under — weeks 1-4 (regular)",
                    "2021-25: 175-144-1, 54.9%, p=.093, 5/5 seasons"),
    "MID_TOTAL_UNDER": ("Under — total 41.5-44.5",
                        "2021-25: 253-214-4, 54.2%, p=.079, 4/5 seasons"),
}


def fetch_games() -> pd.DataFrame:
    r = requests.get(GAMES_URL, timeout=60)
    r.raise_for_status()
    LIVE_CSV.write_bytes(r.content)
    raw = pd.read_csv(LIVE_CSV, low_memory=False)
    df = raw[raw.season == SEASON].copy()
    # same conventions as analyze_nfl.load(): home-perspective spread,
    # negative = home favored; primetime = Mon/Thu or Sun >= 20:00 ET
    df["spread_close"] = -df.spread_line
    hour = pd.to_numeric(df.gametime.astype(str).str.slice(0, 2),
                         errors="coerce")
    df["primetime"] = (df.weekday.isin(["Monday", "Thursday"])
                       | ((df.weekday == "Sunday") & (hour >= 20)))
    return df


def hits(g) -> list[tuple[str, str]]:
    """(rule, bet-text) pairs a game's current lines qualify for."""
    out = []
    sp, tot = g.spread_close, g.total_line
    if pd.notna(tot):
        if g.primetime:
            out.append(("PRIME_UNDER", f"UNDER {tot:g}"))
        if g.game_type == "REG" and g.week <= 4:
            out.append(("EARLY_UNDER", f"UNDER {tot:g}"))
        if 41.5 <= tot <= 44.5:
            out.append(("MID_TOTAL_UNDER", f"UNDER {tot:g}"))
    if pd.notna(sp) and 7 <= abs(sp) <= 9.5:
        dog = g.home_team if sp > 0 else g.away_team
        out.append(("DOG_7_95", f"{dog} +{abs(sp):g}"))
    return out


def grade(rule: str, g) -> str:
    if rule == "DOG_7_95":
        hcm = g.result + g.spread_close        # >0 = home covered
        dog_home = g.spread_close > 0
        if hcm == 0:
            return "P"
        return "W" if (hcm > 0) == dog_home else "L"
    diff = g.total_line - g.total              # >0 = under wins
    return "P" if diff == 0 else ("W" if diff > 0 else "L")


def main():
    df = fetch_games()
    log = json.load(open(LOG, encoding="utf-8")) if LOG.exists() else {}
    today = f"{datetime.date.today():%Y-%m-%d}"

    for g in df.itertuples():
        done = pd.notna(g.result)
        for rule, bet in hits(g):
            key = f"{g.game_id}:{rule}"
            e = log.setdefault(key, dict(
                first_seen=today, rule=rule, week=int(g.week),
                date=g.gameday, home=g.home_team, away=g.away_team,
                first_line=(g.spread_close if rule == "DOG_7_95"
                            else g.total_line)))
            if done and "result" not in e:
                e["close_line"] = (g.spread_close if rule == "DOG_7_95"
                                   else g.total_line)
                e["bet"] = bet
                e["final"] = f"{int(g.home_score)}-{int(g.away_score)}"
                e["result"] = grade(rule, g)
        if done:  # provisional entries that no longer qualify at close
            live = {r for r, _ in hits(g)}
            for rule in RULES:
                key = f"{g.game_id}:{rule}"
                if key in log and rule not in live and "result" not in log[key]:
                    log[key]["dropped"] = "line moved out of bucket by close"
    json.dump(log, open(LOG, "w", encoding="utf-8"), indent=0)

    # ---- scorecard + workbook tab ----
    wb = openpyxl.load_workbook(BOOK)
    if "Watch List 2026" in wb.sheetnames:
        del wb["Watch List 2026"]
    ws = wb.create_sheet("Watch List 2026")
    ws["A1"] = "Pre-registered 2026 paper bets — NFL_WATCHLIST_2026_PREREG.md"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws["A2"] = ("Graded vs nflverse CLOSING lines (eligibility too), pushes "
                "excluded, -110 assumed. Graduation: 52.38% on 100+ decided "
                "bets from 2026 wk 1 — most rules need 2027. No mid-season "
                "edits. first_line is informational only, never graded.")
    ws["A2"].font = Font(name="Arial", italic=True, size=9)

    hdr = Font(name="Arial", bold=True)
    for j, h in enumerate(["Rule", "2021-25 prior", "2026 record",
                           "Win%", "Status"], 1):
        ws.cell(row=4, column=j, value=h).font = hdr
    counts = {}
    for i, (rule, (label, prior)) in enumerate(RULES.items()):
        graded = [e["result"] for e in log.values()
                  if e["rule"] == rule and "result" in e]
        w, l, p = (graded.count(x) for x in "WLP")
        n = w + l
        pct = 100 * w / n if n else None
        pend = sum(1 for e in log.values() if e["rule"] == rule
                   and "result" not in e and "dropped" not in e)
        counts[rule] = (f"{w}-{l}-{p}", pend)
        status = ("GRADUATING" if n >= 100 and pct > 52.38 else
                  "RETIRING — below coin flip" if n >= 100 and pct <= 50 else
                  f"monitoring ({pend} pending)")
        for j, v in enumerate([label, prior, f"{w}-{l}-{p}",
                               f"{pct:.1f}%" if pct is not None else "—",
                               status], 1):
            ws.cell(row=5 + i, column=j, value=v)

    row = 11
    for j, h in enumerate(["Rule", "Wk", "Date", "Matchup", "Bet (close)",
                           "First-seen line", "Final", "Result"], 1):
        ws.cell(row=row, column=j, value=h).font = hdr
    order = {r: i for i, r in enumerate(RULES)}
    entries = sorted(log.values(),
                     key=lambda e: (order[e["rule"]], e["week"]))
    for e in entries:
        row += 1
        res = ("dropped" if "dropped" in e else
               {"W": "Win", "L": "Loss", "P": "Push"}.get(e.get("result"),
                                                          "pending"))
        for j, v in enumerate(
                [RULES[e["rule"]][0], e["week"], e["date"],
                 f"{e['away']} at {e['home']}", e.get("bet", "(pending)"),
                 e["first_line"], e.get("final", ""), res], 1):
            ws.cell(row=row, column=j, value=v)
    for col, w in zip("ABCDEFGH", (30, 5, 12, 30, 16, 14, 10, 10)):
        ws.column_dimensions[col].width = w
    wb.save(BOOK)

    print(f"{SEASON} games in feed: {len(df)}; ledger entries: {len(log)}")
    for rule, (rec, pend) in counts.items():
        print(f"  {rule}: {rec} graded, {pend} pending")


if __name__ == "__main__":
    main()
