"""One-button refresh for NCAA_FBS_Teams: rosters + portal + FPI sheet.

Pipeline:
  1. Re-pull all 138 official rosters + CFBD portal feed (rosters/scripts_2026)
  2. Rebuild FBS_Rosters_2026.xlsx (kept as the raw-roster archive workbook)
  3. Import each team's roster into its tab in NCAA_FBS_Teams (replacing the
     previous roster section - idempotent)
  4. Re-fetch FPI decomposition inputs and rewrite the "FPI Decomposition" sheet

Usage:
  python refresh_all.py                 # full refresh
  python refresh_all.py --import-only   # skip re-fetching; just rebuild sheets
  python refresh_all.py --wait-for-unlock   # used by REFRESH.bat: wait until
                                            # Excel releases the workbook first

Reads CFBD_API_KEY from the user environment (never printed or logged).
"""

from __future__ import annotations

import argparse
import datetime
import os
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
ROSTERS = HERE / "rosters"
FPI_DIR = HERE / "fpi-decomposition"
sys.path.insert(0, str(FPI_DIR))

TEAM_BOOK_XLSM = HERE / "NCAA_FBS_Teams.xlsm"
TEAM_BOOK_XLSX = HERE / "NCAA_FBS_Teams.xlsx"
ROSTER_BOOK = ROSTERS / "FBS_Rosters_2026.xlsx"

ARIAL = Font(name="Arial")
ARIAL_B = Font(name="Arial", bold=True)
HEAD_FILL = PatternFill("solid", start_color="F47321")
WHITE_B = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", start_color="0A2851")
TRANSFER_FILL = PatternFill("solid", start_color="E2EFDA")
NOTE_FILL = PatternFill("solid", start_color="FFF2CC")

ROSTER_HEADERS = ["#", "Name", "Pos", "Class", "Ht", "Wt", "Hometown", "Notes"]
SECTION_AFTER_ROSTER = "Team Stats"  # roster section ends where this begins


def team_book() -> Path:
    return TEAM_BOOK_XLSM if TEAM_BOOK_XLSM.exists() else TEAM_BOOK_XLSX


def load_env_key():
    if not os.environ.get("CFBD_API_KEY"):
        try:
            import winreg
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, "Environment") as k:
                os.environ["CFBD_API_KEY"] = winreg.QueryValueEx(k, "CFBD_API_KEY")[0]
        except OSError:
            pass


def wait_for_unlock(path: Path, timeout=300):
    """Wait until Excel releases the file (button flow closes it first)."""
    start = time.time()
    while time.time() - start < timeout:
        try:
            os.rename(path, path)  # fails on Windows while file is open
            return True
        except OSError:
            time.sleep(2)
    raise TimeoutError(f"{path.name} still locked after {timeout}s - close it in Excel")


def run_roster_refresh():
    for script in ("refresh_2026.py", "build_official.py"):
        p = ROSTERS / "scripts_2026" / script
        print(f"\n== {script} ==")
        r = subprocess.run([sys.executable, str(p)], cwd=str(ROSTERS))
        if r.returncode != 0:
            raise RuntimeError(f"{script} failed (exit {r.returncode})")


# ---------------- roster import ----------------

def read_all_players() -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(ROSTER_BOOK, read_only=True, data_only=True)
    ws = wb["All Players"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    teams: dict[str, list[dict]] = {}
    for conference, team, jersey, name, pos, cls, ht, wt, home, note in rows:
        if not team:
            continue
        teams.setdefault(team, []).append(
            dict(jersey=jersey, name=name, pos=pos, cls=cls, ht=ht, wt=wt,
                 home=home, note=note or ""))
    wb.close()
    return teams


def tab_for_team(sheetnames: list[str], team: str):
    from name_mapping import normalize_name
    # Miami disambiguation: CFBD 'Miami' is the FL school
    special = {"Miami": "Miami (FL)"}
    if team in special and special[team] in sheetnames:
        return special[team]
    by_key = {normalize_name(s): s for s in sheetnames}
    return by_key.get(normalize_name(team))


def find_roster_bounds(ws):
    """Return (start_row, end_row) of the section to replace."""
    start = end = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        v = str(c.value or "")
        if start is None and (v.startswith("2026 Roster") or v.startswith("Depth Chart")):
            start = c.row
        elif start is not None and v.startswith(SECTION_AFTER_ROSTER):
            end = c.row - 1  # keep the blank row before next section? no: replace up to here-1
            break
    if start is None:
        start = ws.max_row + 2
    if end is None:
        end = start + 1 if str(ws.cell(row=start + 1, column=1).value or "").startswith("Section to be") else start
    # swallow trailing blank rows inside the section
    while end > start and ws.cell(row=end, column=1).value in (None, ""):
        end -= 1
    return start, end


def write_team_roster(ws, players: list[dict], stamp: str):
    from openpyxl.worksheet.cell_range import CellRange

    start, end = find_roster_bounds(ws)
    old_n = end - start + 1
    new_n = 2 + len(players) + 1  # header + col headers + players, + trailing blank
    delta = new_n - old_n

    # openpyxl's insert/delete_rows shift cell values but NOT merged ranges,
    # which would leave stale merges masking the roster cells. Unmerge anything
    # at/below the section; re-merge (shifted) only the ranges below it.
    remerge = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= start:
            ws.unmerge_cells(str(mr))
            if mr.min_row > end:
                remerge.append(CellRange(min_col=mr.min_col, max_col=mr.max_col,
                                         min_row=mr.min_row + delta,
                                         max_row=mr.max_row + delta))

    if end >= start:
        ws.delete_rows(start, old_n)
    ws.insert_rows(start, new_n)
    for cr in remerge:
        ws.merge_cells(str(cr))

    ws.cell(row=start, column=1,
            value=f"2026 Roster — official athletics site (refreshed {stamp})").font = ARIAL_B
    for i, h in enumerate(ROSTER_HEADERS, 1):
        c = ws.cell(row=start + 1, column=i, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center")
    for j, p in enumerate(players):
        r = start + 2 + j
        vals = [p["jersey"], p["name"], p["pos"], p["cls"], p["ht"], p["wt"],
                p["home"], p["note"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = ARIAL
        note = str(p["note"] or "")
        if note.startswith(("Transfer", "Incoming")):
            for ci in range(1, 9):
                ws.cell(row=r, column=ci).fill = TRANSFER_FILL
        elif note:
            for ci in range(1, 9):
                ws.cell(row=r, column=ci).fill = NOTE_FILL
    widths = [6, 26, 8, 9, 7, 7, 30, 34]
    for i, w in enumerate(widths, 1):
        col = ws.column_dimensions[get_column_letter(i)]
        if (col.width or 0) < w:
            col.width = w


def load_book(book: Path):
    # keep_vba preserves the vbaProject (REFRESH macro) on .xlsm round-trips
    return openpyxl.load_workbook(book, keep_vba=book.suffix.lower() == ".xlsm")


def import_rosters(book: Path) -> tuple[int, list, list]:
    stamp = f"{datetime.date.today():%B %d, %Y}"
    teams = read_all_players()
    wb = load_book(book)
    done, no_tab = 0, []
    for team, players in sorted(teams.items()):
        tab = tab_for_team(wb.sheetnames, team)
        if not tab:
            no_tab.append(team)
            continue
        write_team_roster(wb[tab], players, stamp)
        done += 1
    skip = {"Overview", "FPI Decomposition"}
    mapped_tabs = {tab_for_team(wb.sheetnames, t) for t in teams}
    empty_tabs = [s for s in wb.sheetnames if s not in skip and s not in mapped_tabs]
    wb.save(book)
    return done, no_tab, empty_tabs


# ---------------- FPI sheet ----------------

def write_fpi_sheet(book: Path, refresh: bool):
    from analysis import FEATURE_COLS, build_dataset, fit_ols

    merged, coverage, year = build_dataset(2025, refresh=refresh, transfer=True)
    model, fitted = fit_ols(merged, FEATURE_COLS)
    model_p, _ = fit_ols(merged, FEATURE_COLS + ["net_portal"])

    t = fitted[["team", "conference", "fpi", "predicted", "residual",
                "prior_sp_raw", "returning_prod_raw", "talent_raw",
                "recruiting_4yr_raw"]].copy()
    t = t.sort_values("fpi", ascending=False).reset_index(drop=True)
    t.insert(0, "fpi_rank", t.index + 1)
    t["resid_rank"] = t["residual"].rank(ascending=False).astype(int)

    wb = load_book(book)
    if "FPI Decomposition" in wb.sheetnames:
        del wb["FPI Decomposition"]
    ws = wb.create_sheet("FPI Decomposition", 1)

    ws["A1"] = "FPI Decomposition — 2025 (ESPN FPI vs public-input model)"
    ws["A1"].font = WHITE_B
    ws["A1"].fill = TITLE_FILL
    ws["A2"] = (f"Source: CFBD API. OLS on {int(model.nobs)} teams: R-sq {model.rsquared:.3f}, "
                f"adj {model.rsquared_adj:.3f}. Residual = FPI minus predicted; positive = ESPN rates the team "
                "higher than public inputs explain. CFBD mirrors final 2025 FPI, not preseason. "
                f"Generated {datetime.date.today():%Y-%m-%d} - see fpi-decomposition/README.md.")
    ws["A2"].font = Font(name="Arial", italic=True, size=9)

    headers = ["FPI Rank", "Team", "Conf", "FPI", "Predicted", "Residual",
               "Resid Rank", "Prior SP+", "Ret PPA", "Talent", "Recruiting 4yr"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = WHITE_B
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center")

    POS = Font(name="Arial", color="006100")
    NEG = Font(name="Arial", color="9C0006")
    for r, row in enumerate(t.itertuples(index=False), start=5):
        vals = [row.fpi_rank, row.team, row.conference, row.fpi, row.predicted,
                row.residual, row.resid_rank, row.prior_sp_raw,
                row.returning_prod_raw, row.talent_raw, row.recruiting_4yr_raw]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=round(v, 2) if isinstance(v, float) else v)
            c.font = ARIAL
            if ci in (4, 5, 6, 8, 9, 10, 11):
                c.number_format = "0.0"
        ws.cell(row=r, column=6).font = POS if row.residual >= 0 else NEG

    for i, w in enumerate([9, 22, 18, 8, 10, 9, 11, 10, 9, 9, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{4 + len(t)}"

    stats = [("Model", "std coef", "p-value")]
    for name in FEATURE_COLS:
        stats.append((name, round(model.params[name], 2),
                      "<0.0001" if model.pvalues[name] < 1e-4 else f"{model.pvalues[name]:.3f}"))
    stats += [
        ("net portal (if added)", round(model_p.params["net_portal"], 2),
         f"{model_p.pvalues['net_portal']:.3f}"),
        ("", "", ""),
        ("Adj R-sq baseline", round(model.rsquared_adj, 4), ""),
        ("Adj R-sq + portal", round(model_p.rsquared_adj, 4), ""),
    ]
    for ri, tup in enumerate(stats, start=4):
        for ci, v in enumerate(tup, start=13):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = WHITE_B if ri == 4 else ARIAL
            if ri == 4:
                c.fill = HEAD_FILL
    wb.save(book)
    print(f"FPI sheet rewritten ({int(model.nobs)} teams, adj R-sq {model.rsquared_adj:.3f})")


def snapshot_preseason_fpi():
    """Once ESPN publishes 2026 preseason FPI and CFBD mirrors it, capture a
    dated snapshot on the first refresh that sees it (before week 1 kicks off,
    this IS the preseason vintage - the true target for the decomposition)."""
    import json
    import cfbd_client as cfbd
    snap_dir = FPI_DIR / "data"
    existing = list(snap_dir.glob("fpi_2026_preseason_snapshot_*.json"))
    if existing:
        return
    try:
        rows = cfbd.get("/ratings/fpi", {"year": 2026}, refresh=True)
    except Exception as exc:
        print(f"WARN: 2026 FPI check failed ({exc}); will retry next refresh")
        return
    if rows:
        path = snap_dir / f"fpi_2026_preseason_snapshot_{datetime.date.today():%Y%m%d}.json"
        json.dump(rows, open(path, "w", encoding="utf-8"))
        print(f"*** 2026 FPI is live on CFBD - preseason snapshot saved: {path.name} ***")
    else:
        print("2026 FPI not published yet - snapshot will be captured automatically "
              "on the first refresh after ESPN releases it.")


def recalc_com(book: Path):
    """Recalculate + save via Excel so formula results are cached in the file."""
    ps = (
        '$xl = New-Object -ComObject Excel.Application; $xl.Visible = $false; '
        '$xl.DisplayAlerts = $false; '
        'try { $wb = $xl.Workbooks.Open("' + str(book) + '"); '
        '$xl.CalculateFullRebuild(); $wb.Save(); $wb.Close($false) } '
        'finally { $xl.Quit() }'
    )
    r = subprocess.run(["powershell", "-NoProfile", "-Command", ps])
    if r.returncode != 0:
        print("WARN: Excel recalc step failed - formulas will compute when the "
              "workbook is opened in Excel.")


# ---------------- main ----------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--import-only", action="store_true",
                    help="skip re-fetching rosters/FPI; just rebuild the sheets from current data")
    ap.add_argument("--wait-for-unlock", action="store_true",
                    help="wait for Excel to release the workbook before starting")
    args = ap.parse_args()

    load_env_key()
    book = team_book()
    print(f"target workbook: {book.name}")

    if args.wait_for_unlock:
        print("waiting for the workbook to be closed in Excel...")
        wait_for_unlock(book)

    if not args.import_only:
        run_roster_refresh()
        snapshot_preseason_fpi()

    print("\n== rebuilding conference tabs + data sheets ==")
    from build_conference_book import restructure
    restructure(book, refresh=not args.import_only)

    print("\n== FPI decomposition sheet ==")
    write_fpi_sheet(book, refresh=False)  # restructure already refreshed the cache

    print("\n== recalculating via Excel ==")
    recalc_com(book)

    print("\nrefresh complete.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
